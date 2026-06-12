# accounts/company_services.py là một module chứa các dịch vụ liên quan đến công ty, bao gồm việc tạo công ty mới, quản lý người dùng và thành viên trong công ty, cũng như xử lý các tác vụ liên quan đến tổ chức và quản lý nhân sự. Các dịch vụ này có thể bao gồm việc tạo tài khoản người dùng, gán vai trò và quyền hạn, quản lý phòng ban và chức vụ, cũng như xử lý các tác vụ liên quan đến việc nhập khẩu dữ liệu nhân sự từ các nguồn khác nhau. Mục tiêu của module này là cung cấp các chức năng cần thiết để quản lý hiệu quả các công ty và nhân sự của họ trong hệ thống.
'''
company_services.py là lớp nghiệp vụ dùng để tạo công ty, tạo nhân viên và nhập dữ liệu công ty từ Excel.

  Nếu models.py trả lời:
  > Dữ liệu công ty được lưu như thế nào?

  thì company_services.py trả lời:
  > Làm thế nào để tạo đầy đủ một công ty cùng tài khoản, nhóm và nhân sự một cách hợp lệ?
'''


from __future__ import annotations

import io
import secrets
import unicodedata
from dataclasses import dataclass
from typing import BinaryIO, Optional

from django.contrib.auth.models import User
from django.db import transaction
from django.utils.text import slugify
# đây là code để import các hàm và lớp từ accounts/identity_normalization.py, bao gồm strip_accents, normalize_lookup_value, normalize_employee_code và build_initials. Các hàm này được sử dụng để chuẩn hóa các giá trị nhận dạng như tên người dùng, mã nhân viên, v.v. trong quá trình quản lý công ty và nhân sự.
from .models import (
    Company,
    CompanyAIConfig,
    CompanyImportBatch,
    CompanyPosition,
    CompanyRole,
    CompanyStatus,
    CompanyUserMembership,
    Department,
    DepartmentMembership,
    UserGroup,
    UserGroupMembership,
    UserProfile,
)

# @dataclass là một decorator trong Python được sử dụng để tự động tạo ra các phương thức đặc biệt như __init__, __repr__, __eq__, v.v. cho một lớp dựa trên các trường dữ liệu được định nghĩa trong lớp đó. frozen=True có nghĩa là các trường dữ liệu của lớp sẽ không thể thay đổi sau khi đối tượng đã được tạo ra, giúp đảm bảo tính bất biến của đối tượng. Trong đoạn code này, BootstrapAdminResult, CompanyCredentialRow và CompanyCreationResult đều là các lớp dữ liệu được định nghĩa bằng cách sử dụng @dataclass để lưu trữ thông tin liên quan đến kết quả của việc tạo công ty và quản lý người dùng trong công ty.
@dataclass(frozen=True)
#class BootstrapAdminResult để lưu trữ kết quả của việc tạo tài khoản quản trị viên cho công ty, bao gồm thông tin về người dùng, thành viên công ty và mật khẩu gốc.
class BootstrapAdminResult:
    user: User
    membership: CompanyUserMembership
    raw_password: str


@dataclass(frozen=True)
#class CompanyCredentialRow để lưu trữ thông tin về một hàng dữ liệu nhân viên trong quá trình tạo công ty, bao gồm tên đầy đủ, email, tên người dùng, mật khẩu, vai trò, phòng ban, chức vụ và mã nhân viên.
# vd: khi tạo một công ty mới, CompanyCredentialRow sẽ chứa thông tin về từng nhân viên được tạo ra, bao gồm tên đầy đủ, email, tên người dùng, mật khẩu, vai trò trong công ty, phòng ban mà họ thuộc về, chức vụ của họ và mã nhân viên nếu có. Thông tin này có thể được sử dụng để xuất ra các báo cáo hoặc để quản lý nhân sự trong công ty sau này.
class CompanyCredentialRow:
    full_name: str
    email: str
    username: str
    password: str
    role: str
    department: str = ''
    position: str = ''
    employee_code: str = ''


@dataclass(frozen=True)
#class CompanyCreationResult để lưu trữ kết quả của việc tạo công ty, bao gồm thông tin về công ty, quản trị viên bootstrap, và các thống kê về các đối tượng đã được tạo.
# vd: khi tạo một công ty mới, CompanyCreationResult sẽ chứa thông tin về công ty đã được tạo, tài khoản quản trị viên đã được bootstrap, số lượng phòng ban, chức vụ và nhân viên đã được tạo, cũng như các hàng dữ liệu nhân viên đã được chuẩn bị để xuất ra hoặc sử dụng trong các bước tiếp theo của quá trình quản lý công ty.
class CompanyCreationResult:
    company: Company
    bootstrap_admin: BootstrapAdminResult
    created_department_count: int
    created_position_count: int
    created_employee_count: int
    credential_rows: tuple[CompanyCredentialRow, ...]
    # Nghiep vu moi: cong ty to chuc theo NHOM (khong dung phong ban/chuc vu).
    created_group_count: int = 0

# def normalize_text để chuẩn hóa một chuỗi văn bản bằng cách loại bỏ các khoảng trắng thừa và chuẩn hóa các ký tự đặc biệt. Nó sử dụng str.strip() để loại bỏ khoảng trắng ở đầu và cuối chuỗi, sau đó sử dụng str.split() để tách chuỗi thành các phần tử dựa trên khoảng trắng, và cuối cùng sử dụng ' '.join() để nối lại các phần tử thành một chuỗi duy nhất với một khoảng trắng giữa chúng. Kết quả trả về là một chuỗi đã được chuẩn hóa, giúp đảm bảo rằng các giá trị văn bản được lưu trữ và so sánh một cách nhất quán trong hệ thống.
# vd: "  Nguyễn   Văn A  " sẽ được chuẩn hóa thành "Nguyễn Văn A" sau khi loại bỏ các khoảng trắng thừa và chuẩn hóa các ký tự đặc biệt, giúp đảm bảo rằng tên nhân viên được lưu trữ và so sánh một cách nhất quán trong hệ thống.
def normalize_text(value) -> str:
    return ' '.join(str(value or '').strip().split())

# def normalize_lookup để chuẩn hóa một giá trị đầu vào cho mục đích tra cứu bằng cách loại bỏ dấu, chuyển đổi thành chữ thường, loại bỏ các ký tự không mong muốn và chuẩn hóa khoảng trắng. Nó sử dụng hàm normalize_text để chuẩn hóa giá trị đầu vào, sau đó sử dụng unicodedata.normalize để loại bỏ dấu và ký tự kết hợp, và cuối cùng loại bỏ các ký tự kết hợp để trả về một chuỗi đã được chuẩn hóa sẵn sàng cho việc tra cứu. Kết quả trả về là một chuỗi đã được chuẩn hóa, giúp đảm bảo rằng các giá trị nhận dạng được chuẩn hóa một cách nhất quán cho mục đích tra cứu trong cơ sở dữ liệu hoặc khi làm việc với chúng trong mã nguồn.
# vd: "  Nguyễn   Văn A  " sẽ được chuẩn hóa thành "Nguyễn Văn A" sau khi loại bỏ các khoảng trắng thừa và chuẩn hóa các ký tự đặc biệt, giúp đảm bảo rằng tên nhân viên được lưu trữ và so sánh một cách nhất quán trong hệ thống.
# normalize_lookup sẽ sử dụng normalize_text và sử dụng thêm bước chuẩn hóa bằng cách loại bỏ dấu và ký tự kết hợp để đảm bảo rằng các giá trị nhận dạng được chuẩn hóa một cách nhất quán cho mục đích tra cứu trong cơ sở dữ liệu hoặc khi làm việc với chúng trong mã nguồn.
# unicodedata.normalize('NFKD', text) sẽ phân tách các ký tự có dấu thành các ký tự cơ bản và các ký tự kết hợp, sau đó ''.join(ch for ch in text if not unicodedata.combining(ch)) sẽ loại bỏ các ký tự kết hợp để trả về một chuỗi đã được chuẩn hóa, giúp đảm bảo rằng các giá trị nhận dạng được chuẩn hóa một cách nhất quán cho mục đích tra cứu trong cơ sở dữ liệu hoặc khi làm việc với chúng trong mã nguồn.
def normalize_lookup(value) -> str:
    text = normalize_text(value).casefold()
    text = unicodedata.normalize('NFKD', text)
    return ''.join(ch for ch in text if not unicodedata.combining(ch))

# def build_technical_username để xây dựng một tên người dùng kỹ thuật dựa trên mã công ty và tên người dùng cục bộ. Nó sử dụng slugify để tạo một chuỗi an toàn từ mã công ty và tên người dùng cục bộ, sau đó thay thế dấu gạch ngang bằng dấu gạch dưới. Nếu chuỗi kết quả dài hơn 140 ký tự, nó sẽ được cắt ngắn để đảm bảo rằng tên người dùng không vượt quá giới hạn độ dài của hệ thống. Nếu tên người dùng đã tồn tại trong cơ sở dữ liệu, nó sẽ thêm một hậu tố số để tạo ra một tên người dùng duy nhất. Kết quả trả về là một tên người dùng kỹ thuật đã được chuẩn hóa và đảm bảo tính duy nhất trong hệ thống.
# vd: nếu mã công ty là "CMP001" và tên người dùng cục bộ là "john.doe", thì build_technical_username sẽ tạo ra một tên người dùng kỹ thuật như "cmp_cmp001_john_doe". Nếu tên người dùng này đã tồn tại trong cơ sở dữ liệu, nó sẽ thêm một hậu tố số như "cmp_cmp001_john_doe_1" để đảm bảo tính duy nhất.
def build_technical_username(company_code: str, local_username: str) -> str:
    base = slugify(f'cmp-{company_code}-{local_username}').replace('-', '_')
    technical_username = base[:140] or 'company_user'
    candidate = technical_username
    suffix = 1
    while User.objects.filter(username=candidate).exists():
        suffix_text = f'_{suffix}'
        candidate = f'{technical_username[: max(1, 140 - len(suffix_text))]}{suffix_text}'
        suffix += 1
    return candidate

# def default_local_username để tạo một tên người dùng cục bộ mặc định dựa trên email hoặc tên đầy đủ của người dùng. Nếu email được cung cấp và chứa ký tự '@', phần trước của email sẽ được sử dụng làm cơ sở cho tên người dùng. Nếu không, tên đầy đủ sẽ được chuẩn hóa và sử dụng để tạo tên người dùng bằng cách thay thế các ký tự không hợp lệ bằng dấu gạch dưới. Kết quả trả về là một tên người dùng cục bộ đã được chuẩn hóa, giúp đảm bảo rằng nó phù hợp với các quy tắc đặt tên trong hệ thống và có thể được sử dụng để tạo tài khoản người dùng mới.
# vd: nếu email là "john.doe@example.com", thì default_local_username sẽ tạo ra một tên người dùng cục bộ như "john_doe". Nếu tên đầy đủ là "John Doe", thì default_local_username sẽ tạo ra một tên người dùng cục bộ như "john_doe". Nếu cả email và tên đầy đủ đều không hợp lệ hoặc không được cung cấp, nó sẽ trả về "user" làm tên người dùng cục bộ mặc định.
def default_local_username(*, email: str = '', full_name: str = '') -> str:
    if email and '@' in email:
        base = email.split('@', 1)[0]
    else:
        base = slugify(full_name).replace('-', '_') or 'user'
    base = base.lower().replace('.', '_')
    return base[:150]

# def _split_full_name để tách tên đầy đủ thành họ và tên. Nó sử dụng hàm normalize_text để chuẩn hóa tên đầy đủ, sau đó tách chuỗi đã được chuẩn hóa thành các phần tử dựa trên khoảng trắng. Nếu chỉ có một phần tử, nó sẽ được coi là họ và tên sẽ để trống. Nếu có nhiều phần tử, phần tử cuối cùng sẽ được coi là họ và phần còn lại sẽ được nối lại thành tên. Kết quả trả về là một tuple chứa họ và tên đã được tách ra từ tên đầy đủ.
# vd: nếu tên đầy đủ là "John Doe", thì _split_full_name sẽ trả về ("Doe", "John"). Nếu tên đầy đủ là "John", thì _split_full_name sẽ trả về ("John", ""). Nếu tên đầy đủ là "John Michael Doe", thì _split_full_name sẽ trả về ("Doe", "John Michael").
def _split_full_name(full_name: str):
    full_name = normalize_text(full_name)
    if not full_name:
        return '', ''
    parts = full_name.split()
    if len(parts) == 1:
        return parts[0], ''
    return parts[-1], ' '.join(parts[:-1])

'''
 Hàm build_company_credential_row() có tác dụng gom thông tin đăng nhập của một nhân viên thành một dòng dữ liệu chuẩn.

  Nó không tạo user, không lưu database và không cấp quyền. Nó chỉ lấy dữ liệu đã có rồi đóng gói thành một object CompanyCredentialRow.
vd:
 Kết quả có dạng:

  CompanyCredentialRow(
      full_name="Nguyễn Văn A",
      email="vana@example.com",
      username="nguyenvana",
      password="mat-khau-tam",
      role="company_user",
      department="Hành chính",
      position="Chuyên viên",
      employee_code="NV001",
  )

  Do frozen=True, dữ liệu không thể bị sửa trực tiếp sau khi tạo:

'''
# vd: khi tạo một công ty mới, build_company_credential_row sẽ được sử dụng để xây dựng các hàng dữ liệu nhân viên dựa trên thông tin người dùng và thành viên công ty, cũng như các thông tin liên quan đến phòng ban, chức vụ và mã nhân viên.
# Ví dụ, nếu có một người dùng với tên đầy đủ "John Doe", email "john.doe@example.com", thì build_company_credential_row sẽ tạo ra một hàng dữ liệu nhân viên với thông tin đã được chuẩn hóa. Nếu người dùng này thuộc về một phòng ban "Sales" và có chức vụ "Manager", thì thông tin này cũng sẽ được chuẩn hóa và bao gồm trong hàng dữ liệu nhân viên. Mã nhân viên nếu có cũng sẽ được chuẩn hóa và bao gồm trong hàng dữ liệu nhân viên. Kết quả cuối cùng là một đối tượng CompanyCredentialRow chứa tất cả thông tin đã được chuẩn hóa về nhân viên này, giúp đảm bảo rằng nó phù hợp với các quy tắc lưu trữ và so sánh trong hệ thống.
def build_company_credential_row(
    *,
    user: User,
    membership: CompanyUserMembership,
    raw_password: str,
    department_name: str = '',
    position_name: str = '',
    employee_code: str = '',
) -> CompanyCredentialRow:
    full_name = normalize_text(user.get_full_name()) or membership.local_username
    return CompanyCredentialRow(
        full_name=full_name,
        email=normalize_text(user.email),
        username=membership.local_username,
        password=raw_password,
        role=membership.role,
        department=normalize_text(department_name),
        position=normalize_text(position_name),
        employee_code=normalize_text(employee_code),
    )

# def serialize_company_credential_rows để chuyển đổi một danh sách các hàng dữ liệu nhân viên (CompanyCredentialRow) thành một danh sách các từ điển, trong đó mỗi từ điển chứa thông tin về một nhân viên. Các trường thông tin bao gồm tên đầy đủ, email, tên người dùng, mật khẩu, vai trò, phòng ban, chức vụ và mã nhân viên. Kết quả trả về là một danh sách các từ điển đã được chuẩn hóa, giúp dễ dàng xuất ra hoặc sử dụng trong các bước tiếp theo của quá trình quản lý công ty.
# vd: khi có một danh sách các đối tượng CompanyCredentialRow chứa thông tin về các nhân viên, serialize_company_credential_rows sẽ chuyển đổi chúng thành một danh sách các từ điển, trong đó mỗi từ điển chứa thông tin đã được chuẩn hóa về một nhân viên. Ví dụ, nếu có một đối tượng CompanyCredentialRow với thông tin về một nhân viên có tên đầy đủ "John Doe", email "john.doe@example.com":
# nó sẽ chuyển đổi từ object CompanyCredentialRow này thành một từ điển 
def serialize_company_credential_rows(credential_rows) -> list[dict]:
    return [
        {
            'full_name': row.full_name,
            'email': row.email,
            'username': row.username,
            'password': row.password,
            'role': row.role,
            'department': row.department,
            'position': row.position,
            'employee_code': row.employee_code,
        }
        for row in credential_rows
    ]

# def _ensure_unique_local_username để đảm bảo rằng tên người dùng cục bộ là duy nhất trong công ty. Nó sử dụng hàm normalize_text để chuẩn hóa tên người dùng cục bộ, sau đó kiểm tra xem tên người dùng đã tồn tại trong cơ sở dữ liệu hay chưa. Nếu đã tồn tại, nó sẽ thêm một hậu tố số để tạo ra một tên người dùng duy nhất. Kết quả trả về là một tên người dùng cục bộ đã được chuẩn hóa và đảm bảo tính duy nhất trong công ty.
# vd: nếu tên người dùng cục bộ là "john_doe" và đã tồn tại trong công ty, thì _ensure_unique_local_username sẽ tạo ra một tên người dùng cục bộ như "john_doe_1". Nếu "john_doe_1" cũng đã tồn tại, nó sẽ tiếp tục tạo ra "john_doe_2", và cứ tiếp tục như vậy cho đến khi tìm được một tên người dùng cục bộ duy nhất trong công ty. Nếu tên người dùng cục bộ ban đầu là "john_doe" và chưa tồn tại trong công ty, thì nó sẽ trả về "john_doe" mà không cần thêm hậu tố số.

def _ensure_unique_local_username(company: Company, local_username: str) -> str:
    base = normalize_text(local_username).lower().replace(' ', '_')[:150] or 'user'
    candidate = base
    suffix = 1
    while CompanyUserMembership.objects.filter(company=company, local_username__iexact=candidate).exists():
        suffix_text = f'_{suffix}'
        candidate = f'{base[: max(1, 150 - len(suffix_text))]}{suffix_text}'
        suffix += 1
    return candidate

# def _build_department_map để xây dựng một bản đồ (dictionary) giữa mã hoặc tên phòng ban đã được chuẩn hóa và đối tượng Department tương ứng trong công ty. Nó lặp qua tất cả các phòng ban của công ty, chuẩn hóa mã và tên của mỗi phòng ban bằng cách sử dụng hàm normalize_lookup, và lưu trữ chúng trong một dictionary với khóa là giá trị đã được chuẩn hóa và giá trị là đối tượng Department. Kết quả trả về là một dictionary giúp dễ dàng tra cứu các phòng ban dựa trên mã hoặc tên đã được chuẩn hóa.
# vd: nếu công ty có một phòng ban với mã "HR" và tên "Human Resources", thì _build_department_map sẽ tạo ra một dictionary như sau:
# {
#     'hr': <Department object for HR>,
#     'human resources': <Department object for Human Resources>,
# }
def _build_department_map(company: Company):
    result = {}
    for department in company.departments.all():
        result[normalize_lookup(department.code)] = department
        result[normalize_lookup(department.name)] = department
    return result

# def _build_position_map để xây dựng một bản đồ (dictionary) giữa mã hoặc tên chức vụ đã được chuẩn hóa và đối tượng CompanyPosition tương ứng trong công ty. Nó lặp qua tất cả các chức vụ của công ty, chuẩn hóa mã và tên của mỗi chức vụ bằng cách sử dụng hàm normalize_lookup, và lưu trữ chúng trong một dictionary với khóa là giá trị đã được chuẩn hóa và giá trị là đối tượng CompanyPosition. Kết quả trả về là một dictionary giúp dễ dàng tra cứu các chức vụ dựa trên mã hoặc tên đã được chuẩn hóa.
# vd: nếu công ty có một chức vụ với mã "MGR" và tên "Manager", thì _build_position_map sẽ tạo ra một dictionary như sau:
# {
#     'mgr': <CompanyPosition object for Manager>,
#     'manager': <CompanyPosition object for Manager>,
# }
def _build_position_map(company: Company):
    result = {}
    for position in company.positions.all():
        result[normalize_lookup(position.code)] = position
        result[normalize_lookup(position.name)] = position
    return result

# def _build_group_map để xây dựng một bản đồ (dictionary) giữa tên nhóm đã được chuẩn hóa và đối tượng UserGroup tương ứng trong công ty. Nó lặp qua tất cả các nhóm người dùng của công ty, chuẩn hóa tên của mỗi nhóm bằng cách sử dụng hàm normalize_lookup, và lưu trữ chúng trong một dictionary với khóa là tên nhóm đã được chuẩn hóa và giá trị là đối tượng UserGroup. Kết quả trả về là một dictionary giúp dễ dàng tra cứu các nhóm người dùng dựa trên tên đã được chuẩn hóa.
# vd: nếu công ty có một nhóm người dùng với tên "Sales Team", thì _build_group_map sẽ tạo ra một dictionary như sau:
# {
#     'sales team': <UserGroup object for Sales Team>,

# }

def _build_group_map(company: Company):
    result = {}
    for group in company.user_groups.all():
        result[normalize_lookup(group.name)] = group
    return result


# def _normalize_role để chuẩn hóa vai trò của một nhân viên thành một trong hai giá trị 'leader' hoặc 'member'. Nó sử dụng hàm normalize_lookup để chuẩn hóa giá trị đầu vào, sau đó kiểm tra xem giá trị đã được chuẩn hóa có thuộc vào tập hợp các từ khóa liên quan đến vai trò lãnh đạo hay không. Nếu có, nó trả về UserGroupMembership.ROLE_LEADER, ngược lại nó trả về UserGroupMembership.ROLE_MEMBER. Kết quả trả về là một chuỗi đại diện cho vai trò đã được chuẩn hóa của nhân viên trong công ty.
# vd: nếu giá trị đầu vào là "Leader", "Trưởng nhóm", "Trưởng nhóm", "TN", hoặc "Lead", thì _normalize_role sẽ trả về UserGroupMembership.ROLE_LEADER. Nếu giá trị đầu vào là bất kỳ chuỗi nào khác, nó sẽ trả về UserGroupMembership.ROLE_MEMBER, giúp đảm bảo rằng vai trò của nhân viên được chuẩn hóa một cách nhất quán trong hệ thống.

def _normalize_role(value) -> str:
    """Chuan hoa vai tro -> 'leader' | 'member'."""
    text = normalize_lookup(value)
    if text in {'leader', 'truongnhom', 'truong nhom', 'tn', 'lead'}:
        return UserGroupMembership.ROLE_LEADER
    return UserGroupMembership.ROLE_MEMBER

# def _normalize_employee_groups để chuẩn hóa thông tin về các nhóm mà một nhân viên thuộc về thành một danh sách các từ điển, trong đó mỗi từ điển chứa tên nhóm và vai trò của nhân viên trong nhóm đó. Hàm này hỗ trợ nhiều định dạng đầu vào khác nhau, bao gồm danh sách các từ điển, danh sách các chuỗi, hoặc một chuỗi định dạng Excel. Kết quả trả về là một danh sách các từ điển đã được chuẩn hóa, giúp đảm bảo rằng thông tin về nhóm của nhân viên được lưu trữ và so sánh một cách nhất quán trong hệ thống.
# vd: nếu đầu vào là một chuỗi định dạng Excel như "Nhom A:leader; Nhom B", thì _normalize_employee_groups sẽ trả về một danh sách các từ điển như sau:
# [
#     {'name': 'Nhom A', 'role': 'leader'},
#     {'name': 'Nhom B', 'role': 'member'},
# ]

def _normalize_employee_groups(item: dict) -> list[dict]:
    """Tra ve danh sach {'name', 'role'} cho cac nhom ma nhan vien thuoc ve.

    Ho tro nhieu dinh dang dau vao:
      - list[dict]: [{'group'|'name': 'Nhom A', 'role': 'leader'|'member'}, ...]
      - list[str] : ['Nhom A', 'Nhom B']
      - str (Excel): 'Nhom A:leader; Nhom B' (phan cach ';' , vai tro sau ':')
    """
    raw = item.get('groups')
    if raw is None:
        raw = item.get('nhom') or item.get('group')
    out: list[dict] = []
    if isinstance(raw, str):
        for chunk in raw.replace('\n', ';').split(';'):
            chunk = chunk.strip()
            if not chunk:
                continue
            if ':' in chunk:
                name, role = chunk.split(':', 1)
            else:
                name, role = chunk, ''
            name = normalize_text(name)
            if name:
                out.append({'name': name, 'role': _normalize_role(role)})
    elif isinstance(raw, (list, tuple)):
        for entry in raw:
            if isinstance(entry, dict):
                name = normalize_text(entry.get('group') or entry.get('name'))
                role = _normalize_role(entry.get('role'))
            else:
                name = normalize_text(entry)
                role = UserGroupMembership.ROLE_MEMBER
            if name:
                out.append({'name': name, 'role': role})
    return out

# def _collect_manual_company_payload_errors để kiểm tra tính hợp lệ của dữ liệu đầu vào khi tạo công ty theo nghiệp vụ nhóm (không sử dụng phòng ban/chức vụ). Nó kiểm tra xem mã công ty, tên công ty, danh sách nhóm và danh sách nhân sự có được cung cấp hay không, cũng như kiểm tra tính duy nhất của tên nhóm và mã nhân viên trong dữ liệu đầu vào. Kết quả trả về là một danh sách các lỗi dưới dạng chuỗi, giúp người dùng biết được những vấn đề cần được sửa chữa trong dữ liệu đầu vào trước khi tiến hành tạo công ty.
def _collect_manual_company_payload_errors(
    *,
    company_data: dict,
    groups: list[dict],
    employees: list[dict],
) -> list[str]:
    """Kiem tra payload tao cong ty theo NGHIEP VU NHOM (khong dung phong ban/chuc vu)."""
    errors: list[str] = []
    if not normalize_text(company_data.get('code')):
        errors.append('Cần nhập mã công ty.')
    if not normalize_text(company_data.get('name')):
        errors.append('Cần nhập tên công ty.')
    if not groups:
        errors.append('Cần tạo ít nhất 1 nhóm.')
    if not employees:
        errors.append('Cần tạo ít nhất 1 nhân sự.')

    group_keys: set[str] = set()
    for index, item in enumerate(groups, start=1):
        name = normalize_text(item.get('name'))
        if not name:
            errors.append(f'Nhóm #{index} cần có tên.')
            continue
        name_key = normalize_lookup(name)
        if name_key in group_keys:
            errors.append(f'Tên nhóm "{name}" bị trùng trong form tạo công ty.')
        group_keys.add(name_key)

    employee_code_keys: set[str] = set()
    local_username_keys: set[str] = set()
    for index, item in enumerate(employees, start=1):
        full_name = normalize_text(item.get('full_name') or item.get('name'))
        if not full_name:
            errors.append(f'Nhân sự #{index} cần có họ tên.')
        emp_groups = _normalize_employee_groups(item)
        if not emp_groups:
            errors.append(
                f'Nhân sự "{full_name or index}" cần được gán ít nhất 1 nhóm.'
            )
        for g in emp_groups:
            if normalize_lookup(g['name']) not in group_keys:
                errors.append(
                    f'Nhân sự "{full_name or index}" tham chiếu nhóm "{g["name"]}" không tồn tại.'
                )
        age_years = item.get('age_years')
        if age_years not in (None, ''):
            try:
                int(age_years)
            except (TypeError, ValueError):
                errors.append(f'Nhân sự "{full_name or index}" có tuổi không hợp lệ.')
        employee_code = normalize_text(item.get('employee_code') or item.get('ma_nhan_vien'))
        if employee_code:
            employee_code_key = normalize_lookup(employee_code)
            if employee_code_key in employee_code_keys:
                errors.append(f'Mã nhân viên "{employee_code}" bị trùng trong form tạo công ty.')
            employee_code_keys.add(employee_code_key)
        local_username = normalize_text(item.get('local_username')).lower()
        if local_username:
            local_username_key = normalize_lookup(local_username)
            if local_username_key in local_username_keys:
                errors.append(f'Username nội bộ "{local_username}" bị trùng trong form tạo công ty.')
            local_username_keys.add(local_username_key)

    return errors

# def create_company_user để tạo một tài khoản người dùng mới cho công ty với các thông tin được cung cấp, bao gồm tên đầy đủ, email, mật khẩu, vai trò, phòng ban, chức vụ và mã nhân viên. Nó đảm bảo rằng tên người dùng cục bộ là duy nhất trong công ty và tạo ra một tên người dùng kỹ thuật dựa trên mã công ty và tên người dùng cục bộ. Sau khi tạo tài khoản người dùng, nó cũng tạo một thành viên công ty liên kết với người dùng đó và cập nhật thông tin hồ sơ của người dùng với các dữ liệu liên quan đến công ty. Nếu có phòng ban được chỉ định, nó sẽ thêm người dùng vào phòng ban đó. Kết quả trả về là một đối tượng BootstrapAdminResult chứa thông tin về người dùng đã được tạo, thành viên công ty liên kết và mật khẩu gốc.
# vd: khi tạo một công ty mới, create_company_user sẽ được sử dụng để tạo ra các tài khoản người dùng cho nhân viên của công ty dựa trên thông tin đã được cung cấp. Ví dụ, nếu có một nhân viên với tên đầy đủ "John Doe", email "john.doe@example.com":
# bootstrap_admin_result là một đối tượng BootstrapAdminResult chứa thông tin về người dùng đã được tạo, thành viên công ty liên kết và mật khẩu gốc. Nếu có phòng ban "Sales" được chỉ định, thì người dùng này sẽ được thêm vào phòng ban "Sales". Kết quả cuối cùng là một tài khoản người dùng mới đã được tạo cho nhân viên này, với tất cả các thông tin liên quan đến công ty đã được cập nhật trong hồ sơ của người dùng.

'''
BootstrapAdminResult là một dataclass dùng để đóng gói kết quả tạo hoặc đặt lại tài khoản quản trị viên ban đầu của công ty.

  @dataclass(frozen=True)
  class BootstrapAdminResult:
      user: User
      membership: CompanyUserMembership
      raw_password: str

  ## “Bootstrap Admin” nghĩa là gì?

  - bootstrap: khởi tạo ban đầu.
  - admin: tài khoản quản trị.

  Khi một công ty mới được tạo, hệ thống cần một tài khoản đầu tiên có quyền quản trị công ty:

  Company mới
  → tạo tài khoản admin ban đầu
  → admin đăng nhập
  → admin quản lý nhân viên và cấu hình công ty

  Tài khoản đó được gọi là bootstrap admin.

  BootstrapAdminResult không phải model và không tạo bảng database. Nó chỉ là object Python chứa kết quả của quy trình trên.

nhưng tại sao def create_company_user tạo user thường lại trả về BootstrapAdminResult


• Vì tên BootstrapAdminResult không còn phản ánh đúng phạm vi sử dụng hiện tại.

  Ban đầu, create_company_user() có thể được viết chủ yếu để tạo admin đầu tiên của công ty:

  bootstrap_admin = create_company_user(
      role=CompanyRole.COMPANY_ADMIN,
  )

  Sau đó hàm được mở rộng để tạo cả nhân viên thường:

  created = create_company_user(
      role=CompanyRole.COMPANY_USER,
  )

  Nhưng kiểu kết quả vẫn giữ tên cũ:


'''

def create_company_user(
    *,
    company: Company,
    local_username: str,
    email: str = '',
    password: Optional[str] = None,
    role: str = CompanyRole.COMPANY_USER,
    first_name: str = '',
    last_name: str = '',
    full_name: str = '',
    profile_data: Optional[dict] = None,
    department: Optional[Department] = None,
    actor: Optional[User] = None,
    must_change_password: bool = False,
) -> BootstrapAdminResult:
    if full_name and not first_name and not last_name:
        first_name, last_name = _split_full_name(full_name)
    raw_password = password or secrets.token_urlsafe(10)
    local_username = _ensure_unique_local_username(company, local_username)
    technical_username = build_technical_username(company.code, local_username)
    user = User.objects.create_user(
        username=technical_username,
        email=normalize_text(email),
        password=raw_password,
        first_name=normalize_text(first_name),
        last_name=normalize_text(last_name),
        is_staff=role == CompanyRole.COMPANY_ADMIN,
        is_superuser=False,
    )
    membership = CompanyUserMembership.objects.create(
        company=company,
        user=user,
        local_username=local_username,
        role=role,
        must_change_password=must_change_password,
    )
    profile = user.profile
    profile.company = company
    profile_data = profile_data or {}
    profile.chuc_danh = normalize_text(profile_data.get('chuc_danh'))
    profile.cccd = normalize_text(profile_data.get('cccd'))
    profile.ma_nhan_vien = normalize_text(profile_data.get('ma_nhan_vien'))
    profile.so_yeu_ly_lich = str(profile_data.get('so_yeu_ly_lich') or '').strip()
    profile.so_dien_thoai = normalize_text(profile_data.get('so_dien_thoai'))
    profile.dia_chi = normalize_text(profile_data.get('dia_chi'))
    profile.bio = str(profile_data.get('bio') or '').strip()
    age_years = profile_data.get('age_years')
    if age_years not in (None, ''):
        try:
            profile.age_years = int(age_years)
        except (TypeError, ValueError):
            profile.age_years = None
    profile.save()
    if department is not None:
        DepartmentMembership.objects.get_or_create(
            department=department,
            user=user,
            defaults={'is_active': True},
        )
    CompanyAIConfig.seed_defaults(company, actor=actor)
    return BootstrapAdminResult(user=user, membership=membership, raw_password=raw_password)

# def reset_company_bootstrap_admin để đặt lại tài khoản quản trị viên ban đầu của công ty. Nó tìm kiếm thành viên công ty có vai trò quản trị viên, nếu không tìm thấy thì sẽ tạo một tài khoản quản trị viên mới với tên người dùng "admin". Nếu tìm thấy, nó sẽ đặt lại mật khẩu của tài khoản đó, kích hoạt tài khoản và yêu cầu người dùng đổi mật khẩu khi đăng nhập lần tiếp theo. Kết quả trả về là một đối tượng BootstrapAdminResult chứa thông tin về người dùng đã được đặt lại hoặc tạo mới, thành viên công ty liên kết và mật khẩu gốc.
# vd: khi một công ty đã có tài khoản quản trị viên nhưng quản trị viên đó đã quên mật khẩu, reset_company_bootstrap_admin sẽ được sử dụng để đặt lại mật khẩu cho tài khoản quản trị viên đó. Nếu công ty chưa có tài khoản quản trị viên nào, reset_company_bootstrap_admin sẽ tạo một tài khoản quản trị viên mới với tên người dùng "admin" và mật khẩu ngẫu nhiên, giúp đảm bảo rằng công ty luôn có một tài khoản quản trị viên để quản lý nhân viên và cấu hình công ty.
def reset_company_bootstrap_admin(
    company: Company,
    *,
    actor: Optional[User] = None,
) -> BootstrapAdminResult:
    membership = (
        CompanyUserMembership.objects.select_related('user')
        .filter(company=company, role=CompanyRole.COMPANY_ADMIN)
        .order_by('pk')
        .first()
    )
    if membership is None:
        return create_company_user(
            company=company,
            local_username='admin',
            email=normalize_text(company.email),
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Company Admin',
            actor=actor,
            must_change_password=True,
        )

    raw_password = secrets.token_urlsafe(10)
    membership.user.set_password(raw_password)
    membership.user.is_active = True
    membership.user.is_staff = True
    membership.user.is_superuser = False
    membership.user.save(update_fields=['password', 'is_active', 'is_staff', 'is_superuser'])
    membership.is_active = True
    membership.must_change_password = True
    membership.save(update_fields=['is_active', 'must_change_password'])
    return BootstrapAdminResult(user=membership.user, membership=membership, raw_password=raw_password)

# def create_company_from_payload để tạo một công ty mới dựa trên dữ liệu đầu vào được cung cấp trong payload. Nó kiểm tra tính hợp lệ của dữ liệu đầu vào, tạo công ty, các nhóm người dùng và nhân viên tương ứng, và trả về một đối tượng CompanyCreationResult chứa thông tin về công ty đã được tạo, tài khoản quản trị viên ban đầu, số lượng phòng ban, chức vụ, nhóm và nhân viên đã được tạo, cũng như các hàng dữ liệu nhân viên đã được chuẩn hóa. Kết quả trả về giúp người dùng biết được thông tin chi tiết về công ty mới đã được tạo ra.
# vd: khi có một payload chứa thông tin về một công ty mới, create_company_from_payload sẽ được sử dụng để tạo ra công ty đó cùng với các nhóm người dùng và nhân viên tương ứng dựa trên thông tin đã được cung cấp. Ví dụ, nếu payload chứa thông tin về một công ty với tên "ABC Corp", mã "ABC001", cùng với một nhóm người dùng "Sales Team" và một nhân viên với tên đầy đủ "John Doe", email "john.doe@example.com", create_company_from_payload sẽ tạo ra công ty "ABC Corp" cùng với nhóm "Sales Team" và nhân viên "John Doe".
def create_company_from_payload(payload: dict, *, actor: Optional[User] = None) -> CompanyCreationResult:
    company_data = payload.get('company') or payload
    groups = list(payload.get('groups') or [])
    employees = list(payload.get('employees') or [])
    company_status = company_data.get('status') or CompanyStatus.ACTIVE
    validation_errors = _collect_manual_company_payload_errors(
        company_data=company_data,
        groups=groups,
        employees=employees,
    )
    if validation_errors:
        raise ValueError(validation_errors)

    with transaction.atomic():
        credential_rows = []
        company = Company.objects.create(
            code=normalize_text(company_data.get('code')).lower() or slugify(company_data.get('name') or 'company'),
            slug=normalize_text(company_data.get('slug')),
            name=normalize_text(company_data.get('name')),
            status=company_status,
            description=str(company_data.get('description') or '').strip(),
            industry=normalize_text(company_data.get('industry')),
            address=normalize_text(company_data.get('address')),
            email=normalize_text(company_data.get('email')),
            phone=normalize_text(company_data.get('phone')),
            website=normalize_text(company_data.get('website')),
            company_context=str(company_data.get('company_context') or '').strip(),
            created_by=actor,
            updated_by=actor,
        )
        CompanyAIConfig.seed_defaults(company, actor=actor)

        for item in groups:
            name = normalize_text(item.get('name'))
            if not name:
                continue
            UserGroup.objects.create(
                company=company,
                name=name,
                description=str(item.get('description') or '').strip(),
                created_by=actor,
            )

        group_map = _build_group_map(company)

        bootstrap_admin = create_company_user(
            company=company,
            local_username='admin',
            email=normalize_text(company_data.get('admin_email')),
            password=company_data.get('admin_password') or None,
            role=CompanyRole.COMPANY_ADMIN,
            full_name=normalize_text(company_data.get('admin_full_name')) or 'Company Admin',
            actor=actor,
            must_change_password=True,
        )
        credential_rows.append(
            build_company_credential_row(
                user=bootstrap_admin.user,
                membership=bootstrap_admin.membership,
                raw_password=bootstrap_admin.raw_password,
                position_name='Company Admin',
            )
        )

        created_employee_count = 0
        for item in employees:
            local_username = item.get('local_username') or default_local_username(
                email=str(item.get('email') or ''),
                full_name=str(item.get('full_name') or item.get('name') or ''),
            )
            emp_groups = _normalize_employee_groups(item)
            chuc_danh = normalize_text(
                item.get('chuc_danh') or item.get('position') or item.get('chuc_vu')
            )
            profile_data = {
                'age_years': item.get('age_years'),
                'so_yeu_ly_lich': item.get('profile_text') or item.get('so_yeu_ly_lich'),
                'bio': item.get('profile_text') or item.get('bio'),
                'ma_nhan_vien': item.get('employee_code') or item.get('ma_nhan_vien'),
                'cccd': item.get('cccd'),
                'so_dien_thoai': item.get('phone') or item.get('so_dien_thoai'),
                'dia_chi': item.get('address') or item.get('dia_chi'),
                'chuc_danh': chuc_danh,
            }
            created = create_company_user(
                company=company,
                local_username=local_username,
                email=normalize_text(item.get('email')),
                password=item.get('password') or None,
                role=CompanyRole.COMPANY_USER,
                full_name=normalize_text(item.get('full_name') or item.get('name')),
                first_name=normalize_text(item.get('first_name')),
                last_name=normalize_text(item.get('last_name')),
                profile_data=profile_data,
                actor=actor,
            )
            # Gan nhan vien vao cac NHOM (vai tro theo tung nhom).
            assigned_group_names: list[str] = []
            for g in emp_groups:
                grp = group_map.get(normalize_lookup(g['name']))
                if grp is None:
                    continue
                UserGroupMembership.objects.get_or_create(
                    group=grp,
                    user=created.user,
                    defaults={'role': g['role']},
                )
                assigned_group_names.append(grp.name)
            credential_rows.append(
                build_company_credential_row(
                    user=created.user,
                    membership=created.membership,
                    raw_password=created.raw_password,
                    department_name=', '.join(assigned_group_names),
                    position_name=chuc_danh,
                    employee_code=profile_data.get('ma_nhan_vien') or '',
                )
            )
            created_employee_count += 1

    return CompanyCreationResult(
        company=company,
        bootstrap_admin=bootstrap_admin,
        created_department_count=0,
        created_position_count=0,
        created_group_count=len(groups),
        created_employee_count=created_employee_count,
        credential_rows=tuple(credential_rows),
    )

#def build_company_import_template_bytes để tạo một tệp Excel mẫu dưới dạng bytes, chứa các sheet với thông tin về nhân sự, danh mục và công ty. Sheet "Sheet1-NhanSu" chứa các cột như tên, tuổi, hồ sơ, nhóm, chức danh, email, số điện thoại, địa chỉ, mã nhân viên và CCCD. Sheet "Sheet2-DanhMuc" chứa các cột như loại, mã, tên và mô tả. Sheet "Sheet3-CongTy" chứa các cột như tên công ty, mã công ty, mô tả, lĩnh vực, địa chỉ, email, điện thoại, website và ngữ cảnh công ty. Kết quả trả về là một tệp Excel dưới dạng bytes có thể được sử dụng làm mẫu để nhập dữ liệu công ty.
def build_company_import_template_bytes(*, company: Optional[Company] = None, include_company_sheet: bool = True) -> bytes:
    import openpyxl

    workbook = openpyxl.Workbook()
    staff_sheet = workbook.active
    staff_sheet.title = 'Sheet1-NhanSu'
    staff_sheet.append([
        'Ten',
        'Tuoi',
        'HoSo',
        'Nhom',
        'ChucDanh',
        'Email',
        'SoDienThoai',
        'DiaChi',
        'MaNhanVien',
        'CCCD',
    ])
    # Cot "Nhom": cho phep gan NHIEU nhom, phan cach bang ";", vai tro sau dau ":"
    # (vd: "Hanh Chinh:leader; Ke Toan"). Khong ghi vai tro = thanh vien.
    staff_sheet.append([
        'Nguyen Van A',
        30,
        'Nhan vien chuyen theo doi van ban den va di.',
        'Hanh Chinh:leader; Ke Toan',
        'Chuyen Vien',
        'nguyenvana@example.com',
        '0912345678',
        '123 Duong ABC, Quan 1',
        'NV001',
        '012345678901',
    ])

    catalog_sheet = workbook.create_sheet('Sheet2-DanhMuc')
    catalog_sheet.append(['Loai', 'Ma', 'Ten', 'MoTa'])
    catalog_sheet.append(['group', '', 'Hanh Chinh', 'Nhom hanh chinh'])
    catalog_sheet.append(['group', '', 'Ke Toan', 'Nhom ke toan'])

    if include_company_sheet:
        company_sheet = workbook.create_sheet('Sheet3-CongTy')
        company_sheet.append([
            'TenCongTy',
            'MaCongTy',
            'MoTa',
            'LinhVuc',
            'DiaChi',
            'Email',
            'DienThoai',
            'Website',
            'NguCanhCongTy',
        ])
        company_sheet.append([
            company.name if company else 'Cong ty mau',
            company.code if company else 'cong-ty-mau',
            company.description if company else 'Mo ta cong ty',
            company.industry if company else 'Van phong',
            company.address if company else '123 Duong Mau',
            company.email if company else 'contact@example.com',
            company.phone if company else '0900000000',
            company.website if company else 'https://example.com',
            company.company_context if company else 'Ngu canh cong ty mau de AI suy dien cho user thuoc cong ty nay.',
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()

# def build_company_credentials_workbook_bytes để tạo một tệp Excel dưới dạng bytes chứa thông tin về tài khoản nhân sự của công ty. Sheet "TaiKhoanNhanSu" chứa các cột như họ tên, email, tên người dùng, mật khẩu, vai trò, phòng ban, chức vụ và mã nhân viên. Sheet "ThongTinCongTy" chứa các cột như tên công ty, mã công ty và tổng số tài khoản. Kết quả trả về là một tệp Excel dưới dạng bytes có thể được sử dụng để lưu trữ hoặc xuất thông tin về tài khoản nhân sự của công ty.
def build_company_credentials_workbook_bytes(
    *,
    company_name: str,
    company_code: str,
    credential_rows,
) -> bytes:
    import openpyxl

    rows = list(credential_rows or [])
    workbook = openpyxl.Workbook()
    credentials_sheet = workbook.active
    credentials_sheet.title = 'TaiKhoanNhanSu'
    credentials_sheet.append(
        [
            'HoTen',
            'Email',
            'Username',
            'Password',
            'VaiTro',
            'PhongBan',
            'ChucVu',
            'MaNhanVien',
        ]
    )
    for row in rows:
        if isinstance(row, dict):
            item = row
        else:
            item = {
                'full_name': row.full_name,
                'email': row.email,
                'username': row.username,
                'password': row.password,
                'role': row.role,
                'department': row.department,
                'position': row.position,
                'employee_code': row.employee_code,
            }
        credentials_sheet.append(
            [
                normalize_text(item.get('full_name')),
                normalize_text(item.get('email')),
                normalize_text(item.get('username')),
                str(item.get('password') or '').strip(),
                normalize_text(item.get('role')),
                normalize_text(item.get('department')),
                normalize_text(item.get('position')),
                normalize_text(item.get('employee_code')),
            ]
        )

    company_sheet = workbook.create_sheet('ThongTinCongTy')
    company_sheet.append(['TenCongTy', normalize_text(company_name)])
    company_sheet.append(['MaCongTy', normalize_text(company_code)])
    company_sheet.append(['TongTaiKhoan', len(rows)])

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_len = 0
            letter = column[0].column_letter
            for cell in column:
                value = '' if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()

# def _sheet_headers để lấy danh sách các tiêu đề cột từ một sheet Excel. Nó sử dụng phương thức iter_rows của sheet để lấy hàng đầu tiên (min_row=1, max_row=1) và chỉ lấy giá trị của các ô (values_only=True). Sau đó, nó chuẩn hóa văn bản của mỗi tiêu đề bằng cách sử dụng hàm normalize_text và trả về một danh sách các tiêu đề đã được chuẩn hóa.
def _sheet_headers(sheet):
    return [normalize_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def _find_header(headers, *candidates):
    normalized_headers = [normalize_lookup(header) for header in headers]
    for candidate in candidates:
        wanted = normalize_lookup(candidate)
        for index, header in enumerate(normalized_headers):
            if header == wanted or wanted in header:
                return index
    return None

# def preview_company_import để xem trước dữ liệu nhập khẩu công ty từ một tệp Excel. Nó kiểm tra sự tồn tại của các sheet cần thiết trong tệp Excel, sau đó trích xuất thông tin về công ty, nhóm và nhân sự từ các sheet tương ứng. Nếu có lỗi trong quá trình trích xuất dữ liệu, nó sẽ lưu trữ các lỗi đó vào một danh sách. Kết quả trả về là một đối tượng CompanyImportBatch chứa thông tin đã được trích xuất và danh sách các lỗi nếu có.
def preview_company_import(excel_file: BinaryIO, *, actor: Optional[User] = None) -> CompanyImportBatch:
    import openpyxl

    workbook = openpyxl.load_workbook(excel_file)
    errors = []

    if 'Sheet1-NhanSu' not in workbook.sheetnames:
        errors.append({'sheet': 'Sheet1-NhanSu', 'row': 0, 'message': 'Thieu sheet nhan su.'})
    if 'Sheet2-DanhMuc' not in workbook.sheetnames:
        errors.append({'sheet': 'Sheet2-DanhMuc', 'row': 0, 'message': 'Thieu sheet danh muc.'})
    if 'Sheet3-CongTy' not in workbook.sheetnames:
        errors.append({'sheet': 'Sheet3-CongTy', 'row': 0, 'message': 'Thieu sheet cong ty.'})

    preview_payload = {
        'company': {},
        'groups': [],
        'employees': [],
    }

    if 'Sheet3-CongTy' in workbook.sheetnames:
        sheet = workbook['Sheet3-CongTy']
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            if len(rows) >= 2 and sum(1 for item in rows[0] if item not in (None, '')) > 1:
                headers = [normalize_text(value) for value in rows[0]]
                values = rows[1] if len(rows) > 1 else []
                lookup = {normalize_lookup(headers[idx]): values[idx] for idx in range(min(len(headers), len(values)))}
            else:
                lookup = {}
                for row in rows:
                    if len(row) >= 2 and row[0]:
                        lookup[normalize_lookup(row[0])] = row[1]
            preview_payload['company'] = {
                'name': normalize_text(lookup.get('tencongty') or lookup.get('ten cong ty')),
                'code': normalize_text(lookup.get('macongty') or lookup.get('ma cong ty')).lower(),
                'description': str(lookup.get('mota') or lookup.get('mo ta') or '').strip(),
                'industry': normalize_text(lookup.get('linhvuc') or lookup.get('nganhnghe') or lookup.get('linh vuc')),
                'address': normalize_text(lookup.get('diachi') or lookup.get('dia chi')),
                'email': normalize_text(lookup.get('email')),
                'phone': normalize_text(lookup.get('dienthoai') or lookup.get('dien thoai')),
                'website': normalize_text(lookup.get('website')),
                'company_context': str(
                    lookup.get('ngucanhcongty')
                    or lookup.get('ngu canh cong ty')
                    or lookup.get('companycontext')
                    or ''
                ).strip(),
                'status': CompanyStatus.ACTIVE,
            }
            if not preview_payload['company']['name']:
                errors.append({'sheet': 'Sheet3-CongTy', 'row': 1, 'message': 'Thieu TenCongTy.'})
            if not preview_payload['company']['code']:
                errors.append({'sheet': 'Sheet3-CongTy', 'row': 1, 'message': 'Thieu MaCongTy.'})

    catalog_group_keys = set()
    if 'Sheet2-DanhMuc' in workbook.sheetnames:
        sheet = workbook['Sheet2-DanhMuc']
        headers = _sheet_headers(sheet)
        # Moi dong la 1 NHOM (cot Ten). Bo qua cot Loai/Ma cu neu co.
        name_idx = _find_header(headers, 'Ten', 'TenNhom', 'Nhom')
        description_idx = _find_header(headers, 'MoTa', 'Mo ta')
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name = normalize_text(row[name_idx] if name_idx is not None and name_idx < len(row) else '')
            description = str(row[description_idx] if description_idx is not None and description_idx < len(row) else '').strip()
            if not name:
                continue
            key = normalize_lookup(name)
            if key in catalog_group_keys:
                continue
            preview_payload['groups'].append({'name': name, 'description': description})
            catalog_group_keys.add(key)

    if 'Sheet1-NhanSu' in workbook.sheetnames:
        sheet = workbook['Sheet1-NhanSu']
        headers = _sheet_headers(sheet)
        name_idx = _find_header(headers, 'Ten', 'HoTen', 'Ho ten')
        age_idx = _find_header(headers, 'Tuoi')
        profile_idx = _find_header(headers, 'HoSo', 'Ho so', 'SoYeuLyLich')
        group_idx = _find_header(headers, 'Nhom', 'Nhom/VaiTro', 'PhongBan', 'Phong ban')
        chuc_danh_idx = _find_header(headers, 'ChucDanh', 'Chuc danh', 'ChucVu', 'Chuc vu')
        email_idx = _find_header(headers, 'Email')
        phone_idx = _find_header(headers, 'SoDienThoai', 'So dien thoai')
        address_idx = _find_header(headers, 'DiaChi', 'Dia chi')
        employee_code_idx = _find_header(headers, 'MaNhanVien', 'Ma nhan vien')
        cccd_idx = _find_header(headers, 'CCCD')

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            full_name = normalize_text(row[name_idx] if name_idx is not None and name_idx < len(row) else '')
            if not full_name and not any(value not in (None, '') for value in row):
                continue
            age_value = row[age_idx] if age_idx is not None and age_idx < len(row) else None
            group_raw = normalize_text(row[group_idx] if group_idx is not None and group_idx < len(row) else '')
            chuc_danh = normalize_text(row[chuc_danh_idx] if chuc_danh_idx is not None and chuc_danh_idx < len(row) else '')
            emp_groups = _normalize_employee_groups({'groups': group_raw})
            if not full_name:
                errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Thieu Ten.'})
            if age_value not in (None, ''):
                try:
                    int(age_value)
                except (TypeError, ValueError):
                    errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Tuoi khong hop le.'})
            if not emp_groups:
                errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Thieu Nhom cho nhan su.'})
            for g in emp_groups:
                if normalize_lookup(g['name']) not in catalog_group_keys:
                    errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': f'Nhom "{g["name"]}" khong co trong Sheet2.'})
            preview_payload['employees'].append(
                {
                    'full_name': full_name,
                    'age_years': age_value,
                    'profile_text': str(row[profile_idx] if profile_idx is not None and profile_idx < len(row) else '').strip(),
                    'groups': emp_groups,
                    'chuc_danh': chuc_danh,
                    'email': normalize_text(row[email_idx] if email_idx is not None and email_idx < len(row) else ''),
                    'phone': normalize_text(row[phone_idx] if phone_idx is not None and phone_idx < len(row) else ''),
                    'address': normalize_text(row[address_idx] if address_idx is not None and address_idx < len(row) else ''),
                    'employee_code': normalize_text(row[employee_code_idx] if employee_code_idx is not None and employee_code_idx < len(row) else ''),
                    'cccd': normalize_text(row[cccd_idx] if cccd_idx is not None and cccd_idx < len(row) else ''),
                }
            )

    batch = CompanyImportBatch.objects.create(
        source_type=CompanyImportBatch.SOURCE_EXCEL,
        status=CompanyImportBatch.STATUS_PREVIEWED if not errors else CompanyImportBatch.STATUS_FAILED,
        uploaded_by=actor,
        preview_payload=preview_payload,
        validation_errors=errors,
        commit_summary={
            'group_count': len(preview_payload['groups']),
            'employee_count': len(preview_payload['employees']),
        },
    )
    return batch

# def commit_company_import để cam kết dữ liệu nhập khẩu công ty đã được xem trước từ một batch. Nó kiểm tra trạng thái của batch và các lỗi xác thực trước khi tiến hành tạo công ty mới dựa trên dữ liệu đã được xem trước. Nếu batch đã được cam kết hoặc có lỗi xác thực, nó sẽ ném ra lỗi. Nếu không, nó sẽ tạo công ty mới bằng cách sử dụng hàm create_company_from_payload và cập nhật thông tin của batch với công ty mục tiêu, trạng thái đã cam kết và tóm tắt kết quả cam kết. Kết quả trả về là một đối tượng CompanyCreationResult chứa thông tin về công ty đã được tạo và các chi tiết liên quan đến quá trình cam kết.
#  commit_company_import() dùng để chính thức tạo công ty và dữ liệu nhân sự từ một bản Excel đã được preview, kiểm tra trước đó.

def commit_company_import(batch: CompanyImportBatch, *, actor: Optional[User] = None) -> CompanyCreationResult:
    if batch.status == CompanyImportBatch.STATUS_COMMITTED:
        raise ValueError('Batch da duoc commit.')
    if batch.validation_errors:
        raise ValueError('Batch preview con loi, khong the commit.')
    result = create_company_from_payload(batch.preview_payload, actor=actor)
    batch.target_company = result.company
    batch.status = CompanyImportBatch.STATUS_COMMITTED
    batch.commit_summary = {
        'company_id': result.company.id,
        'company_code': result.company.code,
        'bootstrap_admin_username': result.bootstrap_admin.membership.local_username,
        'group_count': result.created_group_count,
        'employee_count': result.created_employee_count,
    }
    batch.save(update_fields=['target_company', 'status', 'commit_summary', 'updated_at'])
    return result

# def _find_existing_company_user để tìm kiếm một thành viên công ty hiện có dựa trên các tiêu chí như tên người dùng địa phương, email, mã nhân viên. Nó truy vấn các thành viên công ty và lọc dựa trên các tiêu chí này theo thứ tự ưu tiên: mã nhân viên, email, tên người dùng địa phương. Nếu tìm thấy một thành viên phù hợp, nó sẽ trả về đối tượng CompanyUserMembership tương ứng. Nếu không tìm thấy, nó sẽ trả về None.
def _find_existing_company_user(
    *,
    company: Company,
    local_username: str,
    email: str,
    employee_code: str,
) -> Optional[CompanyUserMembership]:
    memberships = CompanyUserMembership.objects.select_related('user', 'user__profile').filter(company=company)
    if employee_code:
        membership = memberships.filter(user__profile__ma_nhan_vien__iexact=employee_code).first()
        if membership is not None:
            return membership
    if email:
        membership = memberships.filter(user__email__iexact=email).first()
        if membership is not None:
            return membership
    if local_username:
        return memberships.filter(local_username__iexact=local_username).first()
    return None

# def import_company_people_from_excel để nhập khẩu thông tin nhân sự của công ty từ một tệp Excel. Nó kiểm tra sự tồn tại của các sheet cần thiết trong tệp Excel, sau đó trích xuất thông tin về nhóm và nhân sự từ các sheet tương ứng. Nếu có lỗi trong quá trình trích xuất dữ liệu, nó sẽ ném ra lỗi. Kết quả trả về là một dictionary chứa danh sách các nhóm và nhân sự đã được chuẩn hóa từ tệp Excel.
def import_company_people_from_excel(
    excel_file: BinaryIO,
    *,
    company: Company,
    actor: Optional[User] = None,
) -> dict:
    import openpyxl

    workbook = openpyxl.load_workbook(excel_file)
    errors = []
    if 'Sheet1-NhanSu' not in workbook.sheetnames:
        errors.append({'sheet': 'Sheet1-NhanSu', 'row': 0, 'message': 'Thieu sheet nhan su.'})
    if 'Sheet2-DanhMuc' not in workbook.sheetnames:
        errors.append({'sheet': 'Sheet2-DanhMuc', 'row': 0, 'message': 'Thieu sheet danh muc.'})
    if errors:
        raise ValueError(errors)

    group_items = []
    employee_items = []

    # Sheet2-DanhMuc: moi dong la 1 NHOM (cot Ten). (Nghiep vu nhom, bo phong ban/chuc vu.)
    catalog_sheet = workbook['Sheet2-DanhMuc']
    catalog_headers = _sheet_headers(catalog_sheet)
    catalog_name_idx = _find_header(catalog_headers, 'Ten', 'TenNhom', 'Nhom')
    catalog_description_idx = _find_header(catalog_headers, 'MoTa', 'Mo ta')
    group_keys = set()
    for row_number, row in enumerate(catalog_sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = normalize_text(row[catalog_name_idx] if catalog_name_idx is not None and catalog_name_idx < len(row) else '')
        description = str(row[catalog_description_idx] if catalog_description_idx is not None and catalog_description_idx < len(row) else '').strip()
        if not name:
            continue
        key = normalize_lookup(name)
        if key in group_keys:
            continue
        group_items.append({'name': name, 'description': description})
        group_keys.add(key)

    staff_sheet = workbook['Sheet1-NhanSu']
    staff_headers = _sheet_headers(staff_sheet)
    name_idx = _find_header(staff_headers, 'Ten', 'HoTen', 'Ho ten')
    age_idx = _find_header(staff_headers, 'Tuoi')
    profile_idx = _find_header(staff_headers, 'HoSo', 'Ho so', 'SoYeuLyLich')
    group_idx = _find_header(staff_headers, 'Nhom', 'Nhom/VaiTro', 'PhongBan', 'Phong ban')
    chuc_danh_idx = _find_header(staff_headers, 'ChucDanh', 'Chuc danh', 'ChucVu', 'Chuc vu')
    email_idx = _find_header(staff_headers, 'Email')
    phone_idx = _find_header(staff_headers, 'SoDienThoai', 'So dien thoai')
    address_idx = _find_header(staff_headers, 'DiaChi', 'Dia chi')
    employee_code_idx = _find_header(staff_headers, 'MaNhanVien', 'Ma nhan vien')
    cccd_idx = _find_header(staff_headers, 'CCCD')
    seen_employee_codes = set()

    for row_number, row in enumerate(staff_sheet.iter_rows(min_row=2, values_only=True), start=2):
        full_name = normalize_text(row[name_idx] if name_idx is not None and name_idx < len(row) else '')
        if not full_name and not any(value not in (None, '') for value in row):
            continue
        age_value = row[age_idx] if age_idx is not None and age_idx < len(row) else None
        group_raw = normalize_text(row[group_idx] if group_idx is not None and group_idx < len(row) else '')
        chuc_danh = normalize_text(row[chuc_danh_idx] if chuc_danh_idx is not None and chuc_danh_idx < len(row) else '')
        emp_groups = _normalize_employee_groups({'groups': group_raw})
        employee_code = normalize_text(row[employee_code_idx] if employee_code_idx is not None and employee_code_idx < len(row) else '')
        if not full_name:
            errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Thieu Ten.'})
        if age_value not in (None, ''):
            try:
                int(age_value)
            except (TypeError, ValueError):
                errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Tuoi khong hop le.'})
        if not emp_groups:
            errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': 'Thieu Nhom cho nhan su.'})
        for g in emp_groups:
            if normalize_lookup(g['name']) not in group_keys:
                errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': f'Nhom "{g["name"]}" khong co trong Sheet2.'})
        if employee_code:
            lookup_code = normalize_lookup(employee_code)
            if lookup_code in seen_employee_codes:
                errors.append({'sheet': 'Sheet1-NhanSu', 'row': row_number, 'message': f'Ma nhan vien "{employee_code}" bi trung trong file.'})
            seen_employee_codes.add(lookup_code)
        employee_items.append(
            {
                'full_name': full_name,
                'age_years': age_value,
                'profile_text': str(row[profile_idx] if profile_idx is not None and profile_idx < len(row) else '').strip(),
                'groups': emp_groups,
                'chuc_danh': chuc_danh,
                'email': normalize_text(row[email_idx] if email_idx is not None and email_idx < len(row) else ''),
                'phone': normalize_text(row[phone_idx] if phone_idx is not None and phone_idx < len(row) else ''),
                'address': normalize_text(row[address_idx] if address_idx is not None and address_idx < len(row) else ''),
                'employee_code': employee_code,
                'cccd': normalize_text(row[cccd_idx] if cccd_idx is not None and cccd_idx < len(row) else ''),
            }
        )

    if errors:
        raise ValueError(errors)

    results = []
    created_groups = 0
    created_users = 0
    updated_users = 0

    with transaction.atomic():
        for item in group_items:
            group, created = UserGroup.objects.get_or_create(
                company=company,
                name=item['name'],
                defaults={'description': item['description'], 'created_by': actor},
            )
            if not created and item['description'] and group.description != item['description']:
                group.description = item['description']
                group.save(update_fields=['description'])
            created_groups += int(created)

        group_map = _build_group_map(company)

        def _assign_groups(user, emp_groups):
            for g in emp_groups:
                grp = group_map.get(normalize_lookup(g['name']))
                if grp is None:
                    continue
                UserGroupMembership.objects.update_or_create(
                    group=grp,
                    user=user,
                    defaults={'role': g['role']},
                )

        for item in employee_items:
            local_username = default_local_username(
                email=item['email'],
                full_name=item['full_name'],
            )
            membership = _find_existing_company_user(
                company=company,
                local_username=local_username,
                email=item['email'],
                employee_code=item['employee_code'],
            )
            profile_data = {
                'age_years': item.get('age_years'),
                'so_yeu_ly_lich': item.get('profile_text'),
                'bio': item.get('profile_text'),
                'ma_nhan_vien': item.get('employee_code'),
                'cccd': item.get('cccd'),
                'so_dien_thoai': item.get('phone'),
                'dia_chi': item.get('address'),
                'chuc_danh': item.get('chuc_danh'),
            }
            if membership is None:
                created = create_company_user(
                    company=company,
                    local_username=local_username,
                    email=item['email'],
                    role=CompanyRole.COMPANY_USER,
                    full_name=item['full_name'],
                    profile_data=profile_data,
                    actor=actor,
                )
                _assign_groups(created.user, item['groups'])
                created_users += 1
                results.append({
                    'email': created.user.email,
                    'username': created.membership.local_username,
                    'status': 'created',
                })
                continue

            user = membership.user
            first_name, last_name = _split_full_name(item['full_name'])
            if item['email']:
                user.email = item['email']
            if first_name:
                user.first_name = first_name
            if last_name:
                user.last_name = last_name
            user.is_active = True
            user.save(update_fields=['email', 'first_name', 'last_name', 'is_active'])

            profile = user.profile
            profile.company = company
            profile.chuc_danh = profile_data['chuc_danh']
            profile.cccd = profile_data['cccd']
            profile.ma_nhan_vien = profile_data['ma_nhan_vien']
            profile.so_yeu_ly_lich = profile_data['so_yeu_ly_lich']
            profile.so_dien_thoai = profile_data['so_dien_thoai']
            profile.dia_chi = profile_data['dia_chi']
            profile.bio = profile_data['bio']
            try:
                profile.age_years = int(profile_data['age_years']) if profile_data['age_years'] not in (None, '') else None
            except (TypeError, ValueError):
                profile.age_years = None
            profile.save()

            membership.is_active = True
            membership.save(update_fields=['is_active'])
            _assign_groups(user, item['groups'])
            updated_users += 1
            results.append({
                'email': user.email,
                'username': membership.local_username,
                'status': 'updated',
            })

    return {
        'created_groups': created_groups,
        'created_users': created_users,
        'updated_users': updated_users,
        'total': len(employee_items),
        'results': results,
    }

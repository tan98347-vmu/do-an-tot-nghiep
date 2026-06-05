import io
import json

import openpyxl
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from accounts.company_services import create_company_user
from accounts.models import (
    Company,
    CompanyImportBatch,
    CompanyPosition,
    CompanyRole,
    CompanyStatus,
    CompanyUserMembership,
    Department,
    DepartmentMembership,
)


class MultiCompanyAdminApiTests(TestCase):
    def setUp(self):
        self.platform_admin = User.objects.create_superuser(
            username='platform_admin',
            password='platform-pass-123',
            email='platform@example.com',
        )
        self.platform_admin.profile.is_platform_admin_account = True
        self.platform_admin.profile.save(update_fields=['is_platform_admin_account'])
        self.company_a = Company.objects.create(
            code='company-a',
            name='Company A',
            status=CompanyStatus.ACTIVE,
        )
        self.company_b = Company.objects.create(
            code='company-b',
            name='Company B',
            status=CompanyStatus.ACTIVE,
        )
        self.company_a_admin = create_company_user(
            company=self.company_a,
            local_username='admin_a',
            email='admin-a@example.com',
            password='secret12345',
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Admin A',
        )
        self.company_b_admin = create_company_user(
            company=self.company_b,
            local_username='admin_b',
            email='admin-b@example.com',
            password='secret12345',
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Admin B',
        )

    def _build_people_import_file(self):
        workbook = openpyxl.Workbook()
        staff_sheet = workbook.active
        staff_sheet.title = 'Sheet1-NhanSu'
        staff_sheet.append([
            'Ten',
            'Tuoi',
            'HoSo',
            'PhongBan',
            'ChucVu',
            'Email',
            'SoDienThoai',
            'DiaChi',
            'MaNhanVien',
            'CCCD',
        ])
        staff_sheet.append([
            'Nguyen Van C',
            29,
            'Nhan vien xu ly van ban noi bo.',
            'Hanh Chinh',
            'Chuyen Vien',
            'van-c@example.com',
            '0912345678',
            '123 Duong ABC',
            'NV-C',
            '012345678901',
        ])

        catalog_sheet = workbook.create_sheet('Sheet2-DanhMuc')
        catalog_sheet.append(['Loai', 'Ma', 'Ten', 'MoTa'])
        catalog_sheet.append(['department', 'HC', 'Hanh Chinh', 'Phong hanh chinh'])
        catalog_sheet.append(['position', 'CV', 'Chuyen Vien', 'Nhan su nghiep vu'])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'company_people.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _build_company_import_file(self):
        workbook = openpyxl.Workbook()
        staff_sheet = workbook.active
        staff_sheet.title = 'Sheet1-NhanSu'
        staff_sheet.append([
            'Ten',
            'Tuoi',
            'HoSo',
            'PhongBan',
            'ChucVu',
            'Email',
            'SoDienThoai',
            'DiaChi',
            'MaNhanVien',
            'CCCD',
        ])
        staff_sheet.append([
            'Nguyen Van D',
            31,
            'Nhan vien thu nghiem import cong ty.',
            'Hanh Chinh',
            'Chuyen Vien',
            'van-d@example.com',
            '0912345679',
            '456 Duong XYZ',
            'NV-D',
            '012345678912',
        ])

        catalog_sheet = workbook.create_sheet('Sheet2-DanhMuc')
        catalog_sheet.append(['Loai', 'Ma', 'Ten', 'MoTa'])
        catalog_sheet.append(['department', 'HC', 'Hanh Chinh', 'Phong hanh chinh'])
        catalog_sheet.append(['position', 'CV', 'Chuyen Vien', 'Nhan su nghiep vu'])

        company_sheet = workbook.create_sheet('Sheet3-CongTy')
        company_sheet.append([
            'TenCongTy',
            'MaCongTy',
            'MoTa',
            'LinhVuc',
            'DiaChi',
            'Email',
            'DienThoai',
            'Website',
            'NguCanhCongTy',
        ])
        company_sheet.append([
            'Imported Company',
            'imported-company',
            'Cong ty tao tu file Excel.',
            'Van phong',
            '123 Example',
            'contact@imported-company.example.com',
            '02873000000',
            'https://imported-company.example.com',
            'Ngu canh import company.',
        ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'company_import.xlsx',
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_public_company_suggest_returns_only_active_companies(self):
        Company.objects.create(code='company-locked', name='Company Locked', status=CompanyStatus.LOCKED)
        Company.objects.create(code='company-deleted', name='Company Deleted', status=CompanyStatus.DELETED)

        response = self.client.get(reverse('api:public_company_suggest'), {'q': 'company'})

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        returned_codes = {item['code'] for item in payload}
        self.assertIn('company-a', returned_codes)
        self.assertIn('company-b', returned_codes)
        self.assertNotIn('company-locked', returned_codes)
        self.assertNotIn('company-deleted', returned_codes)

    def test_platform_company_list_excludes_deleted_and_trash_lists_deleted(self):
        deleted_company = Company.objects.create(
            code='company-deleted',
            name='Company Deleted',
            status=CompanyStatus.DELETED,
        )
        create_company_user(
            company=deleted_company,
            local_username='admin_deleted',
            email='deleted-admin@example.com',
            password='deleted-pass-123',
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Deleted Admin',
        )
        self.client.force_login(self.platform_admin)

        list_response = self.client.get(reverse('api:platform_company_list_create'))
        trash_response = self.client.get(reverse('api:platform_company_trash'))

        self.assertEqual(list_response.status_code, 200, list_response.content)
        self.assertEqual(trash_response.status_code, 200, trash_response.content)
        listed_codes = {item['code'] for item in list_response.json()}
        trash_codes = {item['code'] for item in trash_response.json()}
        self.assertNotIn('company-deleted', listed_codes)
        self.assertIn('company-deleted', trash_codes)

    def test_platform_admin_can_restore_deleted_company_from_trash(self):
        deleted_company = Company.objects.create(
            code='company-deleted',
            name='Company Deleted',
            status=CompanyStatus.DELETED,
        )
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_restore', kwargs={'pk': deleted_company.pk})
        )

        self.assertEqual(response.status_code, 200, response.content)
        deleted_company.refresh_from_db()
        self.assertEqual(deleted_company.status, CompanyStatus.ACTIVE)

    def test_platform_admin_hard_delete_requires_both_passwords(self):
        deleted_company = Company.objects.create(
            code='company-deleted',
            name='Company Deleted',
            status=CompanyStatus.DELETED,
        )
        deleted_admin = create_company_user(
            company=deleted_company,
            local_username='admin_deleted',
            email='deleted-admin@example.com',
            password='deleted-pass-123',
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Deleted Admin',
        )
        self.client.force_login(self.platform_admin)

        wrong_company_password = self.client.post(
            reverse('api:platform_company_hard_delete', kwargs={'pk': deleted_company.pk}),
            data={
                'company_admin_password': 'wrong-password',
                'platform_admin_password': 'platform-pass-123',
            },
            content_type='application/json',
        )
        wrong_platform_password = self.client.post(
            reverse('api:platform_company_hard_delete', kwargs={'pk': deleted_company.pk}),
            data={
                'company_admin_password': 'deleted-pass-123',
                'platform_admin_password': 'wrong-password',
            },
            content_type='application/json',
        )

        self.assertEqual(wrong_company_password.status_code, 400, wrong_company_password.content)
        self.assertEqual(wrong_platform_password.status_code, 400, wrong_platform_password.content)
        self.assertTrue(Company.objects.filter(pk=deleted_company.pk).exists())
        self.assertTrue(User.objects.filter(pk=deleted_admin.user.pk).exists())

    def test_platform_admin_can_hard_delete_deleted_company(self):
        deleted_company = Company.objects.create(
            code='company-deleted',
            name='Company Deleted',
            status=CompanyStatus.DELETED,
        )
        deleted_admin = create_company_user(
            company=deleted_company,
            local_username='admin_deleted',
            email='deleted-admin@example.com',
            password='deleted-pass-123',
            role=CompanyRole.COMPANY_ADMIN,
            full_name='Deleted Admin',
        )
        deleted_user = create_company_user(
            company=deleted_company,
            local_username='employee_deleted',
            email='deleted-user@example.com',
            password='deleted-pass-456',
            role=CompanyRole.COMPANY_USER,
            full_name='Deleted User',
        )
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_hard_delete', kwargs={'pk': deleted_company.pk}),
            data={
                'company_admin_password': 'deleted-pass-123',
                'platform_admin_password': 'platform-pass-123',
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertFalse(Company.objects.filter(pk=deleted_company.pk).exists())
        self.assertFalse(User.objects.filter(pk=deleted_admin.user.pk).exists())
        self.assertFalse(User.objects.filter(pk=deleted_user.user.pk).exists())

    def test_platform_admin_can_create_company_with_nested_payload(self):
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_list_create'),
            data={
                'code': 'company-c',
                'name': 'Company C',
                'status': CompanyStatus.ACTIVE,
                'company_context': 'Cong ty C phu trach van hanh noi bo.',
                'departments': [
                    {'code': 'HC', 'name': 'Hanh Chinh', 'description': 'Phong hanh chinh'},
                ],
                'positions': [
                    {'code': 'CV', 'name': 'Chuyen Vien', 'description': 'Nhan su nghiep vu'},
                ],
                'employees': [
                    {
                        'full_name': 'Nhan Vien C',
                        'email': 'employee-c@example.com',
                        'department': 'Hanh Chinh',
                        'position': 'Chuyen Vien',
                        'employee_code': 'NV-C-001',
                        'profile_text': 'Ho so nhan vien C.',
                    },
                ],
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201, response.content)
        created_company = Company.objects.get(code='company-c')
        self.assertEqual(created_company.departments.count(), 1)
        self.assertEqual(created_company.positions.count(), 1)
        self.assertEqual(created_company.memberships.count(), 2)
        self.assertTrue(
            CompanyUserMembership.objects.filter(
                company=created_company,
                role=CompanyRole.COMPANY_ADMIN,
                local_username='admin',
            ).exists()
        )
        payload = response.json()
        self.assertEqual(len(payload['credential_rows']), 2)
        self.assertEqual(payload['credential_rows'][0]['username'], 'admin')
        self.assertTrue(payload['credential_rows'][0]['password'])
        self.assertEqual(payload['credential_rows'][1]['email'], 'employee-c@example.com')
        self.assertTrue(payload['credential_rows'][1]['password'])

    def test_platform_manual_company_create_rejects_invalid_nested_payload(self):
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_list_create'),
            data={
                'code': 'company-invalid',
                'name': 'Company Invalid',
                'status': CompanyStatus.ACTIVE,
                'departments': [
                    {'code': 'HC', 'name': 'Hanh Chinh'},
                ],
                'positions': [
                    {'code': 'CV', 'name': 'Chuyen Vien'},
                ],
                'employees': [
                    {
                        'full_name': 'Nhan Vien Loi',
                        'email': 'invalid-employee@example.com',
                        'department': 'Phong Khong Ton Tai',
                        'position': 'Chuyen Vien',
                        'employee_code': 'NV-ERR-001',
                    },
                ],
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400, response.content)
        payload = response.json()
        self.assertIn('errors', payload)
        self.assertIn('Phong Khong Ton Tai', payload['detail'])
        self.assertFalse(Company.objects.filter(code='company-invalid').exists())

    def test_platform_admin_can_reset_bootstrap_admin(self):
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_bootstrap_reset', kwargs={'pk': self.company_a.pk})
        )

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        membership = CompanyUserMembership.objects.get(
            company=self.company_a,
            local_username=payload['username'],
        )
        self.assertTrue(membership.must_change_password)
        self.assertTrue(membership.user.check_password(payload['password']))

    def test_platform_company_detail_returns_bootstrap_and_recent_batches(self):
        self.client.force_login(self.platform_admin)
        preview = {
            'company': {'code': 'preview-co', 'name': 'Preview Co'},
            'departments': [],
            'positions': [],
            'employees': [],
        }
        CompanyImportBatch.objects.create(
            source_type=CompanyImportBatch.SOURCE_EXCEL,
            status=CompanyImportBatch.STATUS_PREVIEWED,
            uploaded_by=self.platform_admin,
            target_company=self.company_a,
            preview_payload=preview,
            validation_errors=[],
            commit_summary={'department_count': 0, 'position_count': 0, 'employee_count': 0},
        )

        response = self.client.get(
            reverse('api:platform_company_detail', kwargs={'pk': self.company_a.pk})
        )

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload['code'], self.company_a.code)
        self.assertEqual(payload['bootstrap_admin']['username'], 'admin_a')
        self.assertEqual(len(payload['recent_import_batches']), 1)

    def test_platform_admin_can_update_company_ai_config(self):
        self.client.force_login(self.platform_admin)

        response = self.client.patch(
            reverse('api:platform_company_ai_config', kwargs={'pk': self.company_a.pk}),
            data={
                'ai_model': 'qwen3:14b',
                'ocr_model': 'qwen3-vl:7b',
                'image_ocr_model': 'qwen3-vl:235b-cloud',
                'company_context': 'Ngu canh cong ty A da duoc cap nhat.',
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.company_a.refresh_from_db()
        config = self.company_a.ai_config
        self.assertEqual(config.ai_model, 'qwen3:14b')
        self.assertEqual(config.ocr_model, 'qwen3-vl:7b')
        self.assertEqual(config.image_ocr_model, 'qwen3-vl:235b-cloud')
        self.assertEqual(self.company_a.company_context, 'Ngu canh cong ty A da duoc cap nhat.')

    def test_platform_import_template_contains_three_official_sheets(self):
        self.client.force_login(self.platform_admin)

        response = self.client.get(reverse('api:platform_company_import_template'))

        self.assertEqual(response.status_code, 200, response.content)
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ['Sheet1-NhanSu', 'Sheet2-DanhMuc', 'Sheet3-CongTy'],
        )

    def test_platform_admin_can_commit_import_and_receive_credentials(self):
        self.client.force_login(self.platform_admin)

        preview_response = self.client.post(
            reverse('api:platform_company_import_preview'),
            data={'excel_file': self._build_company_import_file()},
        )

        self.assertEqual(preview_response.status_code, 201, preview_response.content)
        batch_id = preview_response.json()['batch_id']

        commit_response = self.client.post(
            reverse('api:platform_company_import_commit', kwargs={'batch_id': batch_id})
        )

        self.assertEqual(commit_response.status_code, 200, commit_response.content)
        payload = commit_response.json()
        self.assertEqual(payload['company']['code'], 'imported-company')
        self.assertEqual(len(payload['credential_rows']), 2)
        self.assertEqual(payload['credential_rows'][0]['username'], 'admin')
        self.assertEqual(payload['credential_rows'][1]['email'], 'van-d@example.com')
        self.assertTrue(payload['credential_rows'][1]['password'])

    def test_platform_admin_can_download_credentials_workbook(self):
        self.client.force_login(self.platform_admin)

        response = self.client.post(
            reverse('api:platform_company_credentials_workbook'),
            data=json.dumps({
                'company_name': 'Company Credential Test',
                'company_code': 'company-credential-test',
                'credential_rows': [
                    {
                        'full_name': 'Company Admin',
                        'email': 'admin@example.com',
                        'username': 'admin',
                        'password': 'secret-admin',
                        'role': 'company_admin',
                        'department': '',
                        'position': 'Company Admin',
                        'employee_code': '',
                    },
                    {
                        'full_name': 'Nhan Vien Test',
                        'email': 'employee@example.com',
                        'username': 'employee-test',
                        'password': 'secret-employee',
                        'role': 'company_user',
                        'department': 'Hanh Chinh',
                        'position': 'Chuyen Vien',
                        'employee_code': 'NV-T',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['TaiKhoanNhanSu', 'ThongTinCongTy'])
        credential_sheet = workbook['TaiKhoanNhanSu']
        self.assertEqual(credential_sheet['A2'].value, 'Company Admin')
        self.assertEqual(credential_sheet['C3'].value, 'employee-test')
        self.assertEqual(credential_sheet['D3'].value, 'secret-employee')

    def test_company_admin_department_and_position_endpoints_are_company_scoped(self):
        Department.objects.create(company=self.company_b, name='Phong B', code='PB')
        CompanyPosition.objects.create(company=self.company_b, name='Giam Doc', code='GD')
        self.client.force_login(self.company_a_admin.user)

        create_department = self.client.post(
            reverse('api:admin_departments'),
            data={'code': 'HC', 'name': 'Hanh Chinh', 'description': 'Phong hanh chinh'},
        )
        create_position = self.client.post(
            reverse('api:admin_positions'),
            data={'code': 'CV', 'name': 'Chuyen Vien', 'description': 'Nhan su nghiep vu'},
        )
        list_departments = self.client.get(reverse('api:admin_departments'))
        list_positions = self.client.get(reverse('api:admin_positions'))

        self.assertEqual(create_department.status_code, 201, create_department.content)
        self.assertEqual(create_position.status_code, 201, create_position.content)
        self.assertEqual(list_departments.status_code, 200, list_departments.content)
        self.assertEqual(list_positions.status_code, 200, list_positions.content)
        department_names = {item['name'] for item in list_departments.json()}
        position_names = {item['name'] for item in list_positions.json()}
        self.assertIn('Hanh Chinh', department_names)
        self.assertIn('Chuyen Vien', position_names)
        self.assertNotIn('Phong B', department_names)
        self.assertNotIn('Giam Doc', position_names)

    def test_company_admin_can_update_ai_config(self):
        self.client.force_login(self.company_a_admin.user)

        response = self.client.patch(
            reverse('api:admin_ai_config'),
            data={
                'ai_model': 'deepseek-r1:32b',
                'company_context': 'Ngu canh moi cua company admin.',
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.company_a.refresh_from_db()
        self.assertEqual(self.company_a.ai_config.ai_model, 'deepseek-r1:32b')
        self.assertEqual(self.company_a.company_context, 'Ngu canh moi cua company admin.')

    def test_company_admin_import_template_contains_two_official_sheets(self):
        self.client.force_login(self.company_a_admin.user)

        response = self.client.get(reverse('api:admin_import_template'))

        self.assertEqual(response.status_code, 200, response.content)
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Sheet1-NhanSu', 'Sheet2-DanhMuc'])

    def test_company_admin_import_people_official_excel_template(self):
        self.client.force_login(self.company_a_admin.user)

        response = self.client.post(
            reverse('api:admin_import_users'),
            data={'excel_file': self._build_people_import_file()},
        )

        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload['created_departments'], 1)
        self.assertEqual(payload['created_positions'], 1)
        self.assertEqual(payload['created_users'], 1)
        self.assertEqual(payload['updated_users'], 0)
        membership = CompanyUserMembership.objects.get(
            company=self.company_a,
            user__email='van-c@example.com',
        )
        self.assertEqual(membership.user.email, 'van-c@example.com')
        self.assertEqual(membership.user.profile.ma_nhan_vien, 'NV-C')
        self.assertEqual(membership.user.profile.chuc_danh, 'Chuyen Vien')
        self.assertTrue(
            DepartmentMembership.objects.filter(
                user=membership.user,
                department__company=self.company_a,
                department__name='Hanh Chinh',
                is_active=True,
            ).exists()
        )

    def test_me_password_change_clears_must_change_password(self):
        membership = self.company_a_admin.membership
        membership.must_change_password = True
        membership.save(update_fields=['must_change_password'])
        self.client.force_login(self.company_a_admin.user)

        response = self.client.patch(
            reverse('api:me'),
            data={'password': 'new-secret-123', 'first_name': 'Admin'},
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200, response.content)
        membership.refresh_from_db()
        self.company_a_admin.user.refresh_from_db()
        self.assertFalse(membership.must_change_password)
        self.assertTrue(self.company_a_admin.user.check_password('new-secret-123'))

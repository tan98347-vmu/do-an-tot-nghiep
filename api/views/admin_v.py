"""
Hiện tại, trong admin_v.py, chỉ còn một chức năng chắc chắn được UI hiện tại sử dụng:

  ### Danh sách model Ollama

  GET /api/admin/ollama-models/

  Hàm:

  ollama_models()

  Được màn hình cấu hình AI gọi để:

  - Lấy danh sách model từ Ollama.
  - Phân loại model chat và model embedding.
  - Hiển thị chúng trong dropdown để admin lựa chọn.
"""

import os
import secrets
import datetime
import unicodedata
import io

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from django.conf import settings
from django.contrib.auth.models import User
from django.http import HttpResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404
from django.core.management import call_command
from django.db.utils import DatabaseError, OperationalError, ProgrammingError
from accounts.models import UserGroup, UserGroupMembership, GlobalAIConfig
from accounts.tenancy import is_platform_admin
from ..serializers.admin_s import UserAdminSerializer, UserGroupSerializer, UserGroupDetailSerializer, GlobalAIConfigSerializer, DepartmentSerializer


# Là gì: `_platform_admin_only` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm kiểm tra request hiện tại có quyền quản trị nền tảng hay không; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `is_platform_admin` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu; chuyển kết quả thành HTTP response.
def _platform_admin_only(request):
    if not is_platform_admin(request.user):
        return Response({'detail': 'Chi platform admin moi duoc thao tac.'}, status=status.HTTP_403_FORBIDDEN)
    return None

# Là gì: `_col_idx` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm tìm vị trí cột Excel dựa trên các từ khóa tiêu đề có thể thay đổi; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `str.lower.strip`, `unicodedata.normalize`, `join` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
def _col_idx(headers, *keywords):
    

    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `_col_idx` la helper noi bo cua lop API trong file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Thuong duoc cac ham public nhu `import_users_template`, `import_users_excel`, `user_list` goi lai.
    Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
    """
    # Là gì: `norm` là hàm cục bộ bên trong `_col_idx`, chỉ phục vụ bước xử lý nội bộ của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
    # Chức năng backend: Hàm xử lý phần việc `norm` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
    # Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
    # Mối liên hệ: Hàm phối hợp với `str.lower.strip`, `str.lower`, `unicodedata.normalize` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
    # Bản chất và tác dụng: callback cục bộ chỉ có hiệu lực trong hàm bao ngoài; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
    def norm(s):
        """
        Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
        Vai tro backend: Ham `norm` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
        Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
        Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
        Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
        """
        s = str(s or '').lower().strip()
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(c for c in s if not unicodedata.combining(c))
        return s.replace(' ', '').replace('_', '')
    for kw in keywords:
        kw_n = norm(kw)
        for i, h in enumerate(headers):
            if kw_n in norm(h):
                return i
    return None

# Là gì: `_cell` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm đọc an toàn giá trị một ô trong hàng Excel và áp dụng giá trị mặc định; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `str.strip` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
def _cell(row, idx, default=''):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `_cell` la helper noi bo cua lop API trong file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Thuong duoc cac ham public nhu `import_users_template`, `import_users_excel`, `user_list` goi lai.
    Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
    """
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return str(row[idx]).strip()

# Là gì: `import_users_template` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc dữ liệu đầu vào và chuyển thành bản ghi hệ thống; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `openpyxl.Workbook`, `ws_nhom.append`, `wb.create_sheet` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def import_users_template(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `import_users_template` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem xu ly du lieu hoac thao tac lien quan toi mau van ban theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can xu ly du lieu hoac thao tac lien quan toi mau van ban tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_excel` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac xu ly du lieu hoac thao tac lien quan toi mau van ban tren giao dien.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({'detail': 'Cần cài openpyxl.'}, status=500)

    wb = openpyxl.Workbook()

    
    ws_nhom = wb.active
    ws_nhom.title = 'Nhom'
    nhom_headers = ['ten_nhom', 'mo_ta']
    ws_nhom.append(nhom_headers)
    ws_nhom.append(['Phòng Kỹ thuật', 'Nhóm nhân viên kỹ thuật'])
    ws_nhom.append(['Phòng Kinh doanh', 'Nhóm nhân viên kinh doanh'])

    
    ws_ns = wb.create_sheet('Nhan_Su')
    ns_headers = [
        'email', 'ho_ten', 'role', 'nhom',
        'chuc_danh', 'ma_nhan_vien', 'cccd', 'ngay_sinh',
        'so_yeu_ly_lich',
    ]
    ws_ns.append(ns_headers)
    ws_ns.append([
        'nguyen.van.a@company.com', 'Nguyễn Văn A', 'user', 'Phòng Kỹ thuật',
        'Kỹ sư phần mềm', 'NV001', '001234567890', '01/01/1990',
        'Họ tên: Nguyễn Văn A\nNgày sinh: 01/01/1990\nQuê quán: Hà Nội\nTrình độ: Đại học CNTT',
    ])
    ws_ns.append([
        'tran.thi.b@company.com', 'Trần Thị B', 'leader_gr', 'Phòng Kinh doanh',
        'Trưởng nhóm kinh doanh', 'NV002', '002345678901', '15/06/1988',
        'Họ tên: Trần Thị B\nNgày sinh: 15/06/1988\nChuyên ngành: Quản trị kinh doanh',
    ])

    
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for ws in [ws_nhom, ws_ns]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 20

    
    ws_ns.cell(row=1, column=ns_headers.index('role') + 1).comment = None
    note_row = len(ns_headers) + 3
    ws_ns.cell(row=note_row, column=1).value = (
        'Ghi chú — role: "user" = nhân viên thường | "leader_gr" = trưởng nhóm | "admin" = quản trị viên'
    )
    ws_ns.cell(row=note_row, column=1).font = Font(italic=True, color='666666')

    
    for ws in [ws_nhom, ws_ns]:
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="import_users_template.xlsx"'
    return resp

# Là gì: `import_users_excel` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc dữ liệu đầu vào và chuyển thành bản ghi hệ thống; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `request.FILES.get`, `openpyxl.load_workbook`, `iter_rows` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def import_users_excel(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `import_users_excel` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
    """
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({'detail': 'Cần file excel_file.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        return Response({'detail': f'Không đọc được file Excel: {e}'}, status=status.HTTP_400_BAD_REQUEST)

    log = []
    results = []
    group_obj_map = {}

    
    if 'Nhom' in wb.sheetnames:
        rows = list(wb['Nhom'].iter_rows(min_row=1, values_only=True))
        if rows:
            h = [str(c or '') for c in rows[0]]
            i_ten   = _col_idx(h, 'ten nhom', 'tennhom', 'ten', 'name') or 0
            i_mo_ta = _col_idx(h, 'mo ta', 'mota', 'description') or 1
            for row in rows[1:]:
                ten = _cell(row, i_ten)
                if not ten:
                    continue
                grp, created = UserGroup.objects.get_or_create(
                    name=ten,
                    defaults={'description': _cell(row, i_mo_ta), 'created_by': request.user},
                )
                group_obj_map[ten] = grp
                log.append({'status': 'ok', 'msg': f'{"Tạo" if created else "Cập nhật"} nhóm: {ten}'})

    
    if 'Nhan_Su' not in wb.sheetnames:
        return Response({'detail': 'File Excel phải có sheet "Nhan_Su".'}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(wb['Nhan_Su'].iter_rows(min_row=1, values_only=True))
    if not rows:
        return Response({'detail': 'Sheet "Nhan_Su" trống.'}, status=status.HTTP_400_BAD_REQUEST)

    h = [str(c or '') for c in rows[0]]
    i_email     = _col_idx(h, 'email')
    i_name      = _col_idx(h, 'ho ten', 'hoten', 'ten', 'name', 'ho_ten', 'full')
    i_role      = _col_idx(h, 'role', 'chuc vu', 'vai tro', 'loai')
    i_group     = _col_idx(h, 'nhom', 'group')
    i_chuc_danh = _col_idx(h, 'chuc danh', 'chucdanh', 'chuc_danh', 'position', 'jobtitle')
    i_ma_nv     = _col_idx(h, 'ma nhan vien', 'manhanvien', 'ma_nv', 'manv', 'employee')
    i_cccd      = _col_idx(h, 'cccd', 'cmnd', 'can cuoc')
    i_ngay_sinh = _col_idx(h, 'ngay sinh', 'ngay_sinh', 'dob', 'birth')
    i_so_yeu    = _col_idx(h, 'so yeu ly lich', 'syll', 'so_yeu', 'resume', 'lylich', 'profile')

    
    if i_email is None:
        sample = [list(r) for r in rows[1:4] if r]
        for col in range(len(h)):
            if any('@' in str(r[col] or '') for r in sample if col < len(r)):
                i_email = col
                break
    if i_email is None:
        i_email = 0
    if i_name is None: i_name = 1
    if i_role is None: i_role = 2
    if i_group is None: i_group = 3

    for row_idx, row in enumerate(rows[1:], start=2):
        email = _cell(row, i_email)
        if not email or '@' not in email:
            log.append({'status': 'skip', 'row': row_idx, 'msg': 'Bỏ qua (không có email hợp lệ)'})
            continue

        name        = _cell(row, i_name)
        role_str    = _cell(row, i_role).lower()
        group_name  = _cell(row, i_group)
        chuc_danh   = _cell(row, i_chuc_danh)
        ma_nv       = _cell(row, i_ma_nv)
        cccd        = _cell(row, i_cccd)
        ngay_sinh   = _cell(row, i_ngay_sinh)
        so_yeu      = _cell(row, i_so_yeu)

        
        base_uname = email.split('@')[0].lower().replace('.', '_').replace('-', '_')
        username = base_uname
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f'{base_uname}_{counter}'
            counter += 1

        password = secrets.token_urlsafe(8)
        try:
            parts = name.split(' ', 1)
            user = User.objects.create_user(
                username=username, email=email, password=password,
                first_name=parts[-1] if parts else '',
                last_name=parts[0] if len(parts) > 1 else '',
                is_staff=role_str in ('admin', 'staff'),
                is_superuser=(role_str == 'admin'),
            )

            
            prof_fields = []
            if chuc_danh:
                user.profile.chuc_danh = chuc_danh; prof_fields.append('chuc_danh')
            if ma_nv:
                user.profile.ma_nhan_vien = ma_nv; prof_fields.append('ma_nhan_vien')
            if cccd:
                user.profile.cccd = cccd; prof_fields.append('cccd')
            if so_yeu:
                user.profile.so_yeu_ly_lich = so_yeu; prof_fields.append('so_yeu_ly_lich')
            if ngay_sinh:
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
                    try:
                        user.profile.ngay_sinh = datetime.datetime.strptime(ngay_sinh, fmt).date()
                        prof_fields.append('ngay_sinh')
                        break
                    except ValueError:
                        pass
            if prof_fields:
                user.profile.save(update_fields=prof_fields)

            
            if group_name:
                grp = group_obj_map.get(group_name) or UserGroup.objects.filter(name=group_name).first()
                if grp:
                    mem_role = (UserGroupMembership.ROLE_LEADER
                                if role_str == 'leader_gr' else UserGroupMembership.ROLE_MEMBER)
                    UserGroupMembership.objects.get_or_create(group=grp, user=user, defaults={'role': mem_role})
                else:
                    log.append({'status': 'warn', 'row': row_idx,
                                'msg': f'Nhóm "{group_name}" không tồn tại'})

            log.append({'status': 'ok', 'row': row_idx,
                        'msg': f'Tạo {username} | role={role_str or "user"} | nhóm={group_name or "—"}'})
            results.append({
                'username': username, 'password': password,
                'email': email, 'name': name,
                'group': group_name, 'role': role_str or 'user',
                'chuc_danh': chuc_danh, 'status': 'OK',
            })
        except Exception as e:
            log.append({'status': 'error', 'row': row_idx, 'msg': str(e)})
            results.append({
                'username': username, 'password': '—',
                'email': email, 'name': name,
                'group': group_name, 'role': role_str,
                'chuc_danh': chuc_danh, 'status': f'Lỗi: {e}',
            })

    return Response({
        'imported': len([r for r in results if r['status'] == 'OK']),
        'total': len(results),
        'results': results,
        'log': log,
    })

# Là gì: `user_list` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm truy vấn và trả về danh sách dữ liệu phù hợp; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `User.objects.all.order_by.select_related`, `User.objects.all.order_by`, `UserAdminSerializer` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def user_list(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `user_list` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra danh sach du lieu theo bo loc hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra danh sach du lieu theo bo loc hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra danh sach du lieu theo bo loc hien tai tren giao dien.
    """
    if request.method == 'GET':
        users = User.objects.all().order_by('username').select_related('profile')
        return Response(UserAdminSerializer(users, many=True).data)

    
    data = request.data
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username:
        return Response({'detail': 'Cần nhập tên đăng nhập.'}, status=status.HTTP_400_BAD_REQUEST)
    if not password:
        return Response({'detail': 'Cần nhập mật khẩu.'}, status=status.HTTP_400_BAD_REQUEST)
    if User.objects.filter(username=username).exists():
        return Response({'detail': f'Tên đăng nhập "{username}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        is_superuser = bool(data.get('is_superuser', False))
        is_staff = bool(data.get('is_staff', False)) or is_superuser
        user = User.objects.create_user(
            username=username,
            email=data.get('email', '').strip(),
            password=password,
            first_name=data.get('first_name', '').strip(),
            last_name=data.get('last_name', '').strip(),
            is_staff=is_staff,
            is_superuser=is_superuser,
        )
        chuc_danh = data.get('chuc_danh', '').strip()
        if chuc_danh:
            user.profile.chuc_danh = chuc_danh
            user.profile.save(update_fields=['chuc_danh'])
        return Response(UserAdminSerializer(user).data, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Là gì: `user_detail` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc hoặc xử lý một bản ghi cụ thể; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `get_object_or_404`, `UserAdminSerializer`, `user.delete` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def user_detail(request, pk):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `user_detail` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra du lieu chi tiet cho mot doi tuong cu the theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra du lieu chi tiet cho mot doi tuong cu the tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra du lieu chi tiet cho mot doi tuong cu the tren giao dien.
    """
    user = get_object_or_404(User, pk=pk)

    if request.method == 'GET':
        return Response(UserAdminSerializer(user).data)

    if request.method == 'DELETE':
        if user == request.user:
            return Response({'detail': 'Không thể xóa tài khoản đang đăng nhập.'}, status=status.HTTP_400_BAD_REQUEST)
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    
    data = request.data
    if 'email' in data:
        user.email = data['email']
    if 'first_name' in data:
        user.first_name = data['first_name']
    if 'last_name' in data:
        user.last_name = data['last_name']
    if 'is_active' in data:
        user.is_active = bool(data['is_active'])
    if 'is_superuser' in data:
        user.is_superuser = bool(data['is_superuser'])
        if user.is_superuser:
            user.is_staff = True
    if 'is_staff' in data:
        user.is_staff = bool(data['is_staff']) or user.is_superuser
    if data.get('password'):
        user.set_password(data['password'])
    user.save()
    if 'chuc_danh' in data:
        user.profile.chuc_danh = data['chuc_danh']
        user.profile.save(update_fields=['chuc_danh'])
    return Response(UserAdminSerializer(user).data)

# Là gì: `group_list` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm truy vấn và trả về danh sách dữ liệu phù hợp; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `UserGroup.objects.all`, `UserGroupSerializer`, `request.data.get.strip` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_list(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_list` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra danh sach du lieu theo bo loc hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra danh sach du lieu theo bo loc hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra danh sach du lieu theo bo loc hien tai tren giao dien.
    """
    if request.method == 'GET':
        groups = UserGroup.objects.all()
        return Response(UserGroupSerializer(groups, many=True).data)

    
    name = request.data.get('name', '').strip()
    if not name:
        return Response({'detail': 'Cần nhập tên nhóm.'}, status=status.HTTP_400_BAD_REQUEST)
    if UserGroup.objects.filter(name=name).exists():
        return Response({'detail': f'Nhóm "{name}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
    group = UserGroup.objects.create(
        name=name,
        description=request.data.get('description', '').strip(),
        created_by=request.user,
    )
    return Response(UserGroupSerializer(group).data, status=status.HTTP_201_CREATED)

# Là gì: `group_detail` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc hoặc xử lý một bản ghi cụ thể; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `get_object_or_404`, `UserGroupDetailSerializer`, `group.delete` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_detail(request, pk):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_detail` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra du lieu chi tiet cho mot doi tuong cu the theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra du lieu chi tiet cho mot doi tuong cu the tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra du lieu chi tiet cho mot doi tuong cu the tren giao dien.
    """
    group = get_object_or_404(UserGroup, pk=pk)

    if request.method == 'GET':
        return Response(UserGroupDetailSerializer(group).data)

    if request.method == 'DELETE':
        group.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    
    if 'name' in request.data:
        group.name = request.data['name']
    if 'description' in request.data:
        group.description = request.data['description']
    group.save()
    return Response(UserGroupSerializer(group).data)

# Là gì: `group_members` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm xử lý phần việc `group members` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `get_object_or_404`, `group.memberships.select_related.all`, `group.memberships.select_related` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_members(request, pk):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_members` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
    """
    group = get_object_or_404(UserGroup, pk=pk)

    if request.method == 'GET':
        memberships = group.memberships.select_related('user').all()
        return Response([
            {
                'user_id': m.user.id,
                'username': m.user.username,
                'full_name': f'{m.user.first_name} {m.user.last_name}'.strip() or m.user.username,
                'role': m.role,
            }
            for m in memberships
        ])

    
    user_id = request.data.get('user_id')
    role = request.data.get('role', UserGroupMembership.ROLE_MEMBER)
    if not user_id:
        return Response({'detail': 'Cần user_id.'}, status=status.HTTP_400_BAD_REQUEST)
    if role not in (UserGroupMembership.ROLE_MEMBER, UserGroupMembership.ROLE_LEADER):
        role = UserGroupMembership.ROLE_MEMBER
    member_user = get_object_or_404(User, pk=user_id)
    membership, created = UserGroupMembership.objects.get_or_create(
        group=group, user=member_user, defaults={'role': role}
    )
    if not created:
        membership.role = role
        membership.save()
    return Response({'detail': 'Đã thêm thành viên.'})

# Là gì: `group_member_detail` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc hoặc xử lý một bản ghi cụ thể; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `get_object_or_404`, `membership.delete`, `request.data.get` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_member_detail(request, pk, user_id):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_member_detail` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra du lieu chi tiet cho mot doi tuong cu the theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra du lieu chi tiet cho mot doi tuong cu the tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra du lieu chi tiet cho mot doi tuong cu the tren giao dien.
    """
    group = get_object_or_404(UserGroup, pk=pk)
    membership = get_object_or_404(UserGroupMembership, group=group, user_id=user_id)

    if request.method == 'DELETE':
        membership.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    
    role = request.data.get('role')
    if role not in (UserGroupMembership.ROLE_MEMBER, UserGroupMembership.ROLE_LEADER):
        return Response({'detail': 'role phải là "member" hoặc "leader".'}, status=status.HTTP_400_BAD_REQUEST)
    membership.role = role
    membership.save()
    return Response({'detail': 'Đã cập nhật vai trò.'})

# Là gì: `ollama_models` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm xử lý phần việc `ollama models` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `urllib.request.urlopen`, `_json.loads`, `resp.read` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect lên tệp hoặc storage; chuyển kết quả thành HTTP response.
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def ollama_models(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `ollama_models` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
    """
    import urllib.request
    import json as _json
    from django.conf import settings as django_settings
    try:
        url = f'{django_settings.OLLAMA_BASE_URL}/api/tags'
        with urllib.request.urlopen(url, timeout=4) as resp:
            data = _json.loads(resp.read())
            all_names = [m['name'] for m in data.get('models', [])]
            chat_models = [n for n in all_names if 'embed' not in n.lower()]
            embed_models = [n for n in all_names if 'embed' in n.lower()]
            return Response({'chat_models': chat_models, 'embed_models': embed_models, 'error': None})
    except Exception as e:
        return Response({'chat_models': [], 'embed_models': [], 'error': str(e)})

# Là gì: `ai_config` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm xử lý phần việc `ai config` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `GlobalAIConfig.get_config`, `GlobalAIConfigSerializer`, `serializer.is_valid` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def ai_config(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `ai_config` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
    """
    config = GlobalAIConfig.get_config()
    if request.method == 'GET':
        return Response(GlobalAIConfigSerializer(config).data)
    serializer = GlobalAIConfigSerializer(config, data=request.data, partial=True)
    if serializer.is_valid():
        try:
            serializer.save(updated_by=request.user)
        except (ProgrammingError, OperationalError, DatabaseError) as exc:
            return Response(
                {
                    'detail': f'Database chua cap nhat schema AI config. Hay chay migrate. Chi tiet: {exc}',
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

BACKUP_DIR = os.path.join(settings.BASE_DIR, 'backups')

BACKUP_APPS = {
    'accounts':           ['accounts.userprofile', 'accounts.department', 'accounts.departmentmembership',
                           'accounts.globalaiconfig', 'accounts.usergroup', 'accounts.usergroupmembership'],
    'document_templates': ['document_templates.templatecategory', 'document_templates.documenttemplate',
                           'document_templates.templateversion', 'document_templates.templatepermission',
                           'document_templates.templateaudiencemember',
                           'document_templates.templateapprovallog', 'document_templates.templatefavorite'],
    'documents':          ['documents.documentnumberconfig', 'documents.document'],
    'signing':            ['signing.signingsystemconfig', 'signing.departmentdelegation',
                           'signing.signingproposal', 'signing.signingproposalsigner',
                           'signing.signingpacket', 'signing.signingtask',
                           'signing.signedpdfdocument'],
    'ai_engine':          ['ai_engine.knowledgebase', 'ai_engine.chatsession', 'ai_engine.chatmessage'],
    'prompts':            ['prompts.prompt'],
    'auth':               ['auth.user', 'auth.group', 'auth.permission'],
}

# Là gì: `backup_list` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm truy vấn và trả về danh sách dữ liệu phù hợp; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `_platform_admin_only`, `os.makedirs`, `os.path.isdir` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def backup_list(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `backup_list` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra danh sach du lieu theo bo loc hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra danh sach du lieu theo bo loc hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra danh sach du lieu theo bo loc hien tai tren giao dien.
    """
    platform_error = _platform_admin_only(request)
    if platform_error is not None:
        return platform_error

    os.makedirs(BACKUP_DIR, exist_ok=True)

    
    backup_files = []
    if os.path.isdir(BACKUP_DIR):
        for fname in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if fname.endswith('.json') or fname.endswith('.sql'):
                fpath = os.path.join(BACKUP_DIR, fname)
                stat = os.stat(fpath)
                backup_files.append({
                    'name': fname,
                    'size_kb': stat.st_size // 1024,
                    'modified': datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M:%S'),
                })

    
    from django.apps import apps as django_apps
    db_info = []
    for app_key, labels in BACKUP_APPS.items():
        row = {'app': app_key, 'models': []}
        for label in labels:
            try:
                app_label, model_name = label.split('.')
                model = django_apps.get_model(app_label, model_name)
                count = model.objects.count()
                row['models'].append({'label': label, 'count': count})
            except Exception:
                row['models'].append({'label': label, 'count': '?'})
        db_info.append(row)

    return Response({
        'backup_files': backup_files,
        'db_info': db_info,
        'backup_apps': list(BACKUP_APPS.keys()),
        'backup_dir': BACKUP_DIR,
    })

# Là gì: `backup_create` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm kiểm tra đầu vào và tạo dữ liệu mới; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `_platform_admin_only`, `os.makedirs`, `request.data.get` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect lên tệp hoặc storage; chuyển kết quả thành HTTP response.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def backup_create(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `backup_create` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tao moi ban ghi hoac khoi tao mot luong xu ly theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tao moi ban ghi hoac khoi tao mot luong xu ly tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tao moi ban ghi hoac khoi tao mot luong xu ly tren giao dien.
    """
    platform_error = _platform_admin_only(request)
    if platform_error is not None:
        return platform_error

    os.makedirs(BACKUP_DIR, exist_ok=True)

    backup_type = request.data.get('backup_type', 'full')
    app_key = request.data.get('app_key', '')
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

    try:
        buf = io.StringIO()
        if backup_type == 'full':
            call_command('dumpdata', '--natural-foreign', '--natural-primary',
                         '--indent', '2', stdout=buf)
            fname = f'backup_full_{ts}.json'
        else:
            labels = BACKUP_APPS.get(app_key, [])
            if not labels:
                return Response(
                    {'detail': f'Không tìm thấy app "{app_key}".'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            call_command('dumpdata', *labels, '--natural-foreign', '--natural-primary',
                         '--indent', '2', stdout=buf)
            fname = f'backup_{app_key}_{ts}.json'

        fpath = os.path.join(BACKUP_DIR, fname)
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(buf.getvalue())
        size_kb = os.path.getsize(fpath) // 1024

        return Response({
            'detail': f'Đã tạo backup: {fname} ({size_kb} KB)',
            'filename': fname,
            'size_kb': size_kb,
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'detail': f'Lỗi tạo backup: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Là gì: `backup_download` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm chuẩn bị và trả tệp cho phía client tải xuống; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `_platform_admin_only`, `os.path.join`, `os.path.isfile` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect lên tệp hoặc storage; chuyển kết quả thành HTTP response.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def backup_download(request, filename):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `backup_download` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra tep de frontend tai xuong theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra tep de frontend tai xuong tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra tep de frontend tai xuong tren giao dien.
    """
    platform_error = _platform_admin_only(request)
    if platform_error is not None:
        return platform_error

    if '..' in filename:
        raise Http404
    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.isfile(fpath):
        raise Http404
    return FileResponse(open(fpath, 'rb'), as_attachment=True, filename=filename)

# Là gì: `backup_delete` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm xóa hoặc đánh dấu xóa dữ liệu được chỉ định; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `_platform_admin_only`, `os.path.join`, `os.path.isfile` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def backup_delete(request, filename):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `backup_delete` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem xoa hoac don du lieu khong con hieu luc theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can xoa hoac don du lieu khong con hieu luc tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac xoa hoac don du lieu khong con hieu luc tren giao dien.
    """
    platform_error = _platform_admin_only(request)
    if platform_error is not None:
        return platform_error

    if '..' in filename:
        return Response({'detail': 'Tên file không hợp lệ.'}, status=status.HTTP_400_BAD_REQUEST)
    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.isfile(fpath):
        raise Http404
    os.remove(fpath)
    return Response({'detail': f'Đã xóa backup: {filename}'}, status=status.HTTP_200_OK)

# Là gì: `company_context_read` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc dữ liệu từ nguồn được chỉ định; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `GlobalAIConfig.get_config` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def company_context_read(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `company_context_read` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem chuan bi ngu canh cho buoc xu ly phia sau theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can chuan bi ngu canh cho buoc xu ly phia sau tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac chuan bi ngu canh cho buoc xu ly phia sau tren giao dien.
    """
    config = GlobalAIConfig.get_config()
    return Response({'company_context': config.company_context or ''})

# Là gì: `_normalize_group_name` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm chuẩn hóa dữ liệu về định dạng thống nhất; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `join`, `str.strip.split`, `str.strip` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
def _normalize_group_name(name):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `_normalize_group_name` la helper noi bo cua lop API trong file `api/views/admin_v.py`, chiu trach nhiem chuan hoa du lieu dau vao hoac du lieu trung gian truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can chuan hoa du lieu dau vao hoac du lieu trung gian nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Thuong duoc cac ham public nhu `import_users_template`, `import_users_excel`, `user_list` goi lai.
    Tac dung: Co lap rieng buoc chuan hoa du lieu dau vao hoac du lieu trung gian de cac endpoint cung file tai su dung dung mot quy tac.
    """
    return ' '.join(str(name or '').strip().split())

# Là gì: `_find_duplicate_group` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm xử lý phần việc `find duplicate group` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `_normalize_group_name.casefold`, `_normalize_group_name`, `UserGroup.objects.all` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
def _find_duplicate_group(name, *, exclude_id=None):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `_find_duplicate_group` la helper noi bo cua lop API trong file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Thuong duoc cac ham public nhu `import_users_template`, `import_users_excel`, `user_list` goi lai.
    Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
    """
    normalized = _normalize_group_name(name).casefold()
    if not normalized:
        return None
    qs = UserGroup.objects.all()
    if exclude_id is not None:
        qs = qs.exclude(pk=exclude_id)
    for group in qs.only('id', 'name'):
        if _normalize_group_name(group.name).casefold() == normalized:
            return group
    return None

# Là gì: `_group_usage_counts` là helper nội bộ của module `admin_v.py`, phục vụ nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ.
# Chức năng backend: Hàm xử lý phần việc `group usage counts` theo dữ liệu và ngữ cảnh được truyền vào; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Flutter không gọi trực tiếp hàm này; các endpoint cùng module dùng kết quả của nó để phục vụ các màn hình quản trị hoặc API tương thích cũ.
# Mối liên hệ: Hàm phối hợp với `group.memberships.count`, `Document.objects.filter.count`, `DocumentTemplate.objects.filter.count` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: hàm hỗ trợ tái sử dụng trong module; chủ yếu đọc, kiểm tra hoặc biến đổi dữ liệu.
def _group_usage_counts(group):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `_group_usage_counts` la helper noi bo cua lop API trong file `api/views/admin_v.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Thuong duoc cac ham public nhu `import_users_template`, `import_users_excel`, `user_list` goi lai.
    Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
    """
    from documents.models import Document
    from document_templates.models import DocumentTemplate, PendingTemplateAssignment

    return {
        'members': group.memberships.count(),
        'documents': Document.objects.filter(group=group).count(),
        'templates': DocumentTemplate.objects.filter(group=group).count(),
        'pending_template_assignments': PendingTemplateAssignment.objects.filter(group=group).count(),
    }

# Là gì: `group_list` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm truy vấn và trả về danh sách dữ liệu phù hợp; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `UserGroup.objects.all`, `UserGroupSerializer`, `_normalize_group_name` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_list(request):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_list` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra danh sach du lieu theo bo loc hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra danh sach du lieu theo bo loc hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra danh sach du lieu theo bo loc hien tai tren giao dien.
    """
    if request.method == 'GET':
        groups = UserGroup.objects.all()
        return Response(UserGroupSerializer(groups, many=True).data)

    name = _normalize_group_name(request.data.get('name', ''))
    if not name:
        return Response({'detail': 'Can nhap ten nhom.'}, status=status.HTTP_400_BAD_REQUEST)
    duplicate = _find_duplicate_group(name)
    if duplicate:
        return Response({'detail': f'Nhom "{duplicate.name}" da ton tai.'}, status=status.HTTP_400_BAD_REQUEST)

    group = UserGroup.objects.create(
        name=name,
        description=request.data.get('description', '').strip(),
        created_by=request.user,
    )
    return Response(UserGroupSerializer(group).data, status=status.HTTP_201_CREATED)

# Là gì: `group_detail` là endpoint REST của nhóm quản trị hệ thống, cấu hình AI, import người dùng và backup kiểu cũ; nó là điểm nhận request từ client đã đi qua router và lớp permission.
# Chức năng backend: Hàm đọc hoặc xử lý một bản ghi cụ thể; đầu vào được kiểm tra hoặc chuẩn hóa trước khi tạo kết quả.
# Vai trò với UI: Kết quả được các màn hình quản trị hoặc API tương thích cũ sử dụng trực tiếp để hiển thị dữ liệu, tải tệp hoặc cập nhật trạng thái thao tác.
# Mối liên hệ: Hàm phối hợp với `get_object_or_404`, `UserGroupDetailSerializer`, `group.save` và trả dữ liệu về cho lớp gọi kế tiếp trong cùng luồng.
# Bản chất và tác dụng: view mỏng ở biên HTTP; có side effect ghi cơ sở dữ liệu; chuyển kết quả thành HTTP response.
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def group_detail(request, pk):
    """
    Thuoc chuc nang nao: Bang dieu khien, Tai khoan - phong ban - nhom, Cau hinh AI, Sao luu du lieu, thong bao va ha tang route chung.
    Vai tro backend: Ham `group_detail` la endpoint hoac diem vao REST cua file `api/views/admin_v.py`, chiu trach nhiem tra du lieu chi tiet cho mot doi tuong cu the theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can tra du lieu chi tiet cho mot doi tuong cu the tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `my_tennis_club.settings`, `my_tennis_club.urls`, `api/urls.py`, `api/views/dashboard.py`, `api/views/notifications.py`, `accounts.models`. Dung cung cap voi cac ham `_col_idx`, `_cell`, `import_users_template` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac tra du lieu chi tiet cho mot doi tuong cu the tren giao dien.
    """
    from django.db import transaction
    from documents.models import Document
    from document_templates.models import DocumentTemplate, PendingTemplateAssignment

    group = get_object_or_404(UserGroup, pk=pk)

    if request.method == 'GET':
        return Response(UserGroupDetailSerializer(group).data)

    if request.method == 'PATCH':
        if 'name' in request.data:
            normalized_name = _normalize_group_name(request.data.get('name', ''))
            if not normalized_name:
                return Response({'detail': 'Can nhap ten nhom.'}, status=status.HTTP_400_BAD_REQUEST)
            duplicate = _find_duplicate_group(normalized_name, exclude_id=group.pk)
            if duplicate:
                return Response({'detail': f'Nhom "{duplicate.name}" da ton tai.'}, status=status.HTTP_400_BAD_REQUEST)
            group.name = normalized_name
        if 'description' in request.data:
            group.description = request.data['description']
        group.save()
        return Response(UserGroupSerializer(group).data)

    confirmed = str(
        request.query_params.get('confirm')
        or request.data.get('confirm')
        or ''
    ).strip().lower() in {'1', 'true', 'yes'}
    if not confirmed:
        return Response(
            {
                'detail': 'Can xac nhan truoc khi xoa nhom.',
                'requires_confirmation': True,
                'usage': _group_usage_counts(group),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    destination_group_id = request.query_params.get('destination_group_id') or request.data.get('destination_group_id')
    usage = _group_usage_counts(group)
    requires_destination_group = any(usage.values())
    destination_group = None

    if destination_group_id not in (None, ''):
        destination_group = get_object_or_404(UserGroup, pk=destination_group_id)
        if destination_group.pk == group.pk:
            return Response({'detail': 'Khong the chuyen du lieu vao cung mot nhom.'}, status=status.HTTP_400_BAD_REQUEST)

    if requires_destination_group and destination_group is None:
        return Response(
            {
                'detail': 'Nhom nay dang co thanh vien hoac du lieu lien quan. Hay chon destination_group_id truoc khi xoa.',
                'requires_destination_group': True,
                'usage': usage,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    with transaction.atomic():
        if destination_group is not None:
            UserGroupMembership.objects.filter(group=group).update(group=destination_group)
            Document.objects.filter(group=group).update(group=destination_group)
            DocumentTemplate.objects.filter(group=group).update(group=destination_group)
            PendingTemplateAssignment.objects.filter(group=group).update(group=destination_group)
        group.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

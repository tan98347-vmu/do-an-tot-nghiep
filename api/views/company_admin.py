import secrets

from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.company_services import (
    build_company_import_template_bytes,
    create_company_user,
    default_local_username,
    import_company_people_from_excel,
    normalize_text,
)
from accounts.models import (
    CompanyAIConfig,
    CompanyPosition,
    CompanyRole,
    CompanyUserMembership,
    Department,
    UserGroup,
    UserGroupMembership,
)
from accounts.tenancy import get_user_company, is_company_admin
from ..serializers.company_admin import (
    CompanyAdminGroupDetailSerializer,
    CompanyAdminGroupSerializer,
    CompanyAdminDepartmentSerializer,
    CompanyAdminPositionSerializer,
    CompanyAdminUserSerializer,
    CompanyAIConfigSerializer,
)


def _company_admin_context(request):
    company = get_user_company(request.user)
    if company is None or not is_company_admin(request.user):
        return None, Response({'detail': 'Chi company admin moi duoc thao tac.'}, status=status.HTTP_403_FORBIDDEN)
    return company, None


def _company_users(company):
    return User.objects.filter(company_membership__company=company).select_related('profile', 'company_membership')


def _serialize_user(user, company):
    return CompanyAdminUserSerializer(user, context={'company': company}).data


def _group_queryset(company):
    return UserGroup.objects.filter(company=company).order_by('name')


def _department_queryset(company):
    return Department.objects.filter(company=company).order_by('name', 'code')


def _position_queryset(company):
    return CompanyPosition.objects.filter(company=company).order_by('name', 'code')


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def user_list(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    if request.method == 'GET':
        users = _company_users(company).order_by('company_membership__local_username')
        return Response(CompanyAdminUserSerializer(users, many=True, context={'company': company}).data)

    data = request.data
    role = CompanyRole.COMPANY_ADMIN if bool(data.get('is_staff')) or bool(data.get('is_superuser')) else CompanyRole.COMPANY_USER
    local_username = normalize_text(data.get('username')) or default_local_username(
        email=str(data.get('email') or ''),
        full_name=f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
    )
    result = create_company_user(
        company=company,
        local_username=local_username,
        email=normalize_text(data.get('email')),
        password=str(data.get('password') or '') or None,
        role=role,
        first_name=normalize_text(data.get('first_name')),
        last_name=normalize_text(data.get('last_name')),
        profile_data={'chuc_danh': normalize_text(data.get('chuc_danh'))},
        actor=request.user,
        must_change_password=bool(data.get('must_change_password')),
    )
    if 'is_active' in data and not bool(data.get('is_active')):
        result.user.is_active = False
        result.user.save(update_fields=['is_active'])
        result.membership.is_active = False
        result.membership.save(update_fields=['is_active'])
    return Response(_serialize_user(result.user, company), status=status.HTTP_201_CREATED)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def user_detail(request, pk):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    user = get_object_or_404(_company_users(company), pk=pk)
    membership = user.company_membership

    if request.method == 'GET':
        return Response(_serialize_user(user, company))

    if request.method == 'DELETE':
        if user.pk == request.user.pk:
            return Response({'detail': 'Khong the xoa tai khoan dang dang nhap.'}, status=status.HTTP_400_BAD_REQUEST)
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data
    if 'username' in data:
        new_local_username = normalize_text(data.get('username')).lower()
        if new_local_username and CompanyUserMembership.objects.filter(company=company, local_username__iexact=new_local_username).exclude(pk=membership.pk).exists():
            return Response({'detail': 'Ten dang nhap da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        if new_local_username:
            membership.local_username = new_local_username
    if 'email' in data:
        user.email = normalize_text(data.get('email'))
    if 'first_name' in data:
        user.first_name = normalize_text(data.get('first_name'))
    if 'last_name' in data:
        user.last_name = normalize_text(data.get('last_name'))
    if 'is_active' in data:
        active = bool(data.get('is_active'))
        user.is_active = active
        membership.is_active = active
    if 'is_staff' in data or 'is_superuser' in data:
        membership.role = CompanyRole.COMPANY_ADMIN if bool(data.get('is_staff')) or bool(data.get('is_superuser')) else CompanyRole.COMPANY_USER
        user.is_staff = membership.role == CompanyRole.COMPANY_ADMIN
        user.is_superuser = False
    if data.get('password'):
        user.set_password(str(data.get('password')))
        membership.must_change_password = False
    user.save()
    membership.save()
    if 'chuc_danh' in data:
        user.profile.chuc_danh = normalize_text(data.get('chuc_danh'))
        user.profile.save(update_fields=['chuc_danh'])
    return Response(_serialize_user(user, company))


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def group_list(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    if request.method == 'GET':
        groups = _group_queryset(company)
        return Response(CompanyAdminGroupSerializer(groups, many=True).data)

    name = normalize_text(request.data.get('name'))
    if not name:
        return Response({'detail': 'Can nhap ten nhom.'}, status=status.HTTP_400_BAD_REQUEST)
    if UserGroup.objects.filter(company=company, name__iexact=name).exists():
        return Response({'detail': 'Ten nhom da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
    group = UserGroup.objects.create(
        company=company,
        name=name,
        description=str(request.data.get('description') or '').strip(),
        created_by=request.user,
    )
    return Response(CompanyAdminGroupSerializer(group).data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def group_detail(request, pk):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    group = get_object_or_404(_group_queryset(company), pk=pk)
    if request.method == 'GET':
        return Response(CompanyAdminGroupDetailSerializer(group).data)
    if request.method == 'DELETE':
        group.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    if 'name' in request.data:
        name = normalize_text(request.data.get('name'))
        if not name:
            return Response({'detail': 'Can nhap ten nhom.'}, status=status.HTTP_400_BAD_REQUEST)
        if UserGroup.objects.filter(company=company, name__iexact=name).exclude(pk=group.pk).exists():
            return Response({'detail': 'Ten nhom da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        group.name = name
    if 'description' in request.data:
        group.description = str(request.data.get('description') or '').strip()
    group.save()
    return Response(CompanyAdminGroupSerializer(group).data)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def group_members(request, pk):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    group = get_object_or_404(_group_queryset(company), pk=pk)

    if request.method == 'GET':
        return Response(CompanyAdminGroupDetailSerializer(group).data['members'])

    user = get_object_or_404(_company_users(company), pk=request.data.get('user_id'))
    role = request.data.get('role') or UserGroupMembership.ROLE_MEMBER
    membership, created = UserGroupMembership.objects.get_or_create(
        group=group,
        user=user,
        defaults={'role': role},
    )
    if not created and membership.role != role:
        membership.role = role
        membership.save(update_fields=['role'])
    return Response({'detail': 'Da cap nhat thanh vien nhom.'}, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def group_member_detail(request, pk, user_id):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    group = get_object_or_404(_group_queryset(company), pk=pk)
    membership = get_object_or_404(UserGroupMembership.objects.select_related('user', 'group'), group=group, user_id=user_id)

    if request.method == 'DELETE':
        membership.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    role = request.data.get('role') or UserGroupMembership.ROLE_MEMBER
    membership.role = role
    membership.save(update_fields=['role'])
    return Response({'detail': 'Da cap nhat vai tro thanh vien.'})


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def ai_config(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    config = CompanyAIConfig.seed_defaults(company, actor=request.user)
    if request.method == 'GET':
        return Response(CompanyAIConfigSerializer(config).data)

    serializer = CompanyAIConfigSerializer(config, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save(updated_by=request.user)
    if 'company_context' in serializer.validated_data:
        company.company_context = serializer.validated_data['company_context']
        company.updated_by = request.user
        company.save(update_fields=['company_context', 'updated_by', 'updated_at'])
    return Response(CompanyAIConfigSerializer(config).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def company_context_read(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error
    config = CompanyAIConfig.seed_defaults(company, actor=request.user)
    return Response({'company_context': config.company_context or ''})


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def department_list(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    if request.method == 'GET':
        departments = _department_queryset(company)
        return Response(CompanyAdminDepartmentSerializer(departments, many=True).data)

    name = normalize_text(request.data.get('name'))
    code = normalize_text(request.data.get('code')).upper()
    if not name:
        return Response({'detail': 'Can nhap ten phong ban.'}, status=status.HTTP_400_BAD_REQUEST)
    if not code:
        return Response({'detail': 'Can nhap ma phong ban.'}, status=status.HTTP_400_BAD_REQUEST)
    if Department.objects.filter(company=company, name__iexact=name).exists():
        return Response({'detail': 'Ten phong ban da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
    if Department.objects.filter(company=company, code__iexact=code).exists():
        return Response({'detail': 'Ma phong ban da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
    department = Department.objects.create(
        company=company,
        name=name,
        code=code,
        description=str(request.data.get('description') or '').strip(),
        is_active=bool(request.data.get('is_active', True)),
    )
    return Response(CompanyAdminDepartmentSerializer(department).data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def department_detail(request, pk):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    department = get_object_or_404(_department_queryset(company), pk=pk)
    if request.method == 'GET':
        return Response(CompanyAdminDepartmentSerializer(department).data)
    if request.method == 'DELETE':
        if department.memberships.filter(is_active=True).exists():
            return Response({'detail': 'Phong ban dang co nhan su hoat dong, hay vo hieu hoa thay vi xoa.'}, status=status.HTTP_400_BAD_REQUEST)
        department.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    if 'name' in request.data:
        name = normalize_text(request.data.get('name'))
        if not name:
            return Response({'detail': 'Can nhap ten phong ban.'}, status=status.HTTP_400_BAD_REQUEST)
        if Department.objects.filter(company=company, name__iexact=name).exclude(pk=department.pk).exists():
            return Response({'detail': 'Ten phong ban da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        department.name = name
    if 'code' in request.data:
        code = normalize_text(request.data.get('code')).upper()
        if not code:
            return Response({'detail': 'Can nhap ma phong ban.'}, status=status.HTTP_400_BAD_REQUEST)
        if Department.objects.filter(company=company, code__iexact=code).exclude(pk=department.pk).exists():
            return Response({'detail': 'Ma phong ban da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        department.code = code
    if 'description' in request.data:
        department.description = str(request.data.get('description') or '').strip()
    if 'is_active' in request.data:
        department.is_active = bool(request.data.get('is_active'))
    department.save()
    return Response(CompanyAdminDepartmentSerializer(department).data)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def position_list(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    if request.method == 'GET':
        positions = _position_queryset(company)
        return Response(CompanyAdminPositionSerializer(positions, many=True).data)

    name = normalize_text(request.data.get('name'))
    code = normalize_text(request.data.get('code'))
    if not name:
        return Response({'detail': 'Can nhap ten chuc vu.'}, status=status.HTTP_400_BAD_REQUEST)
    if CompanyPosition.objects.filter(company=company, name__iexact=name).exists():
        return Response({'detail': 'Ten chuc vu da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
    if code and CompanyPosition.objects.filter(company=company, code__iexact=code).exists():
        return Response({'detail': 'Ma chuc vu da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
    position = CompanyPosition.objects.create(
        company=company,
        code=code,
        name=name,
        description=str(request.data.get('description') or '').strip(),
        is_active=bool(request.data.get('is_active', True)),
    )
    return Response(CompanyAdminPositionSerializer(position).data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def position_detail(request, pk):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    position = get_object_or_404(_position_queryset(company), pk=pk)
    if request.method == 'GET':
        return Response(CompanyAdminPositionSerializer(position).data)
    if request.method == 'DELETE':
        if company.profiles.filter(chuc_danh__iexact=position.name).exists():
            return Response({'detail': 'Chuc vu dang duoc gan cho nhan su, hay vo hieu hoa thay vi xoa.'}, status=status.HTTP_400_BAD_REQUEST)
        position.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    old_name = position.name
    if 'name' in request.data:
        name = normalize_text(request.data.get('name'))
        if not name:
            return Response({'detail': 'Can nhap ten chuc vu.'}, status=status.HTTP_400_BAD_REQUEST)
        if CompanyPosition.objects.filter(company=company, name__iexact=name).exclude(pk=position.pk).exists():
            return Response({'detail': 'Ten chuc vu da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        position.name = name
    if 'code' in request.data:
        code = normalize_text(request.data.get('code'))
        if code and CompanyPosition.objects.filter(company=company, code__iexact=code).exclude(pk=position.pk).exists():
            return Response({'detail': 'Ma chuc vu da ton tai trong cong ty.'}, status=status.HTTP_400_BAD_REQUEST)
        position.code = code
    if 'description' in request.data:
        position.description = str(request.data.get('description') or '').strip()
    if 'is_active' in request.data:
        position.is_active = bool(request.data.get('is_active'))
    position.save()
    if old_name != position.name:
        company.profiles.filter(chuc_danh__iexact=old_name).update(chuc_danh=position.name)
    return Response(CompanyAdminPositionSerializer(position).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def import_users_template(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error
    response = HttpResponse(
        build_company_import_template_bytes(company=company, include_company_sheet=False),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="import_company_people_template.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_users_excel(request):
    company, error = _company_admin_context(request)
    if error is not None:
        return error

    excel_file = request.FILES.get('excel_file')
    if excel_file is None:
        return Response({'detail': 'Can file excel_file.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(excel_file)
    except Exception as exc:
        return Response({'detail': f'Khong doc duoc file Excel: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    if {'Sheet1-NhanSu', 'Sheet2-DanhMuc'}.issubset(set(workbook.sheetnames)):
        try:
            summary = import_company_people_from_excel(excel_file, company=company, actor=request.user)
            return Response(summary)
        except ValueError as exc:
            return Response({'detail': 'File Excel khong hop le.', 'errors': list(exc.args[0])}, status=status.HTTP_400_BAD_REQUEST)

    log = []
    results = []
    group_map = {}

    if 'Nhom' in workbook.sheetnames:
        rows = list(workbook['Nhom'].iter_rows(min_row=2, values_only=True))
        for row in rows:
            name = normalize_text(row[0] if len(row) > 0 else '')
            description = str(row[1] if len(row) > 1 and row[1] is not None else '').strip()
            if not name:
                continue
            group, created = UserGroup.objects.get_or_create(
                company=company,
                name=name,
                defaults={'description': description, 'created_by': request.user},
            )
            if not created and description and group.description != description:
                group.description = description
                group.save(update_fields=['description'])
            group_map[name] = group
            log.append({'status': 'ok', 'msg': f'{"Tao" if created else "Dung lai"} nhom: {name}'})

    if 'Nhan_Su' not in workbook.sheetnames:
        return Response({'detail': 'File Excel phai co sheet "Nhan_Su".'}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(workbook['Nhan_Su'].iter_rows(min_row=2, values_only=True))
    for index, row in enumerate(rows, start=2):
        email = normalize_text(row[0] if len(row) > 0 else '')
        full_name = normalize_text(row[1] if len(row) > 1 else '')
        role_text = normalize_text(row[2] if len(row) > 2 else '').lower()
        group_name = normalize_text(row[3] if len(row) > 3 else '')
        chuc_danh = normalize_text(row[4] if len(row) > 4 else '')
        ma_nhan_vien = normalize_text(row[5] if len(row) > 5 else '')
        cccd = normalize_text(row[6] if len(row) > 6 else '')
        ngay_sinh = normalize_text(row[7] if len(row) > 7 else '')
        so_yeu_ly_lich = str(row[8] if len(row) > 8 and row[8] is not None else '').strip()

        if not email or '@' not in email:
            log.append({'status': 'skip', 'row': index, 'msg': 'Bo qua dong khong co email hop le.'})
            continue

        local_username = default_local_username(email=email, full_name=full_name)
        membership_role = CompanyRole.COMPANY_ADMIN if role_text in {'admin', 'staff', 'company_admin'} else CompanyRole.COMPANY_USER
        try:
            created = create_company_user(
                company=company,
                local_username=local_username,
                email=email,
                password=secrets.token_urlsafe(8),
                role=membership_role,
                full_name=full_name or email.split('@', 1)[0],
                profile_data={
                    'chuc_danh': chuc_danh,
                    'ma_nhan_vien': ma_nhan_vien,
                    'cccd': cccd,
                    'so_yeu_ly_lich': so_yeu_ly_lich,
                    'bio': so_yeu_ly_lich,
                },
                actor=request.user,
            )
            if ngay_sinh:
                try:
                    created.user.profile.ngay_sinh = ngay_sinh
                    created.user.profile.save(update_fields=['ngay_sinh'])
                except Exception:
                    pass
            if group_name:
                group = group_map.get(group_name)
                if group is None:
                    group = UserGroup.objects.create(company=company, name=group_name, created_by=request.user)
                    group_map[group_name] = group
                UserGroupMembership.objects.update_or_create(
                    group=group,
                    user=created.user,
                    defaults={
                        'role': UserGroupMembership.ROLE_LEADER if role_text == 'leader_gr' else UserGroupMembership.ROLE_MEMBER,
                    },
                )
            results.append(
                {
                    'email': email,
                    'username': created.membership.local_username,
                    'group': group_name,
                    'role': role_text or 'user',
                    'status': 'OK',
                }
            )
        except Exception as exc:
            results.append(
                {
                    'email': email,
                    'group': group_name,
                    'role': role_text or 'user',
                    'status': f'Loi: {exc}',
                }
            )

    return Response(
        {
            'imported': len([item for item in results if item['status'] == 'OK']),
            'total': len(results),
            'results': results,
            'log': log,
        }
    )

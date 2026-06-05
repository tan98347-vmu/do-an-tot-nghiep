"""
Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
Vai tro backend: File `api/views/bulk_upload.py` giu hoac ho tro luong backend cho CRUD mau, duyet mau, version mau, import DOCX/URL, bulk upload va preview noi dung mau.
Vai tro cua no trong frontend: Cac man `/templates`, `/templates/create`, man chi tiet mau va man bulk upload lay du lieu hoac chiu tac dong gian tiep tu file nay.
Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`.
Tac dung: Giu cho du lieu mau, quyen thao tac va preview cua nhom man Mau van ban luon dong nhat giua API, storage va chi so tim kiem.
"""
import re as _re
import json as _json
import io
import unicodedata

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from document_templates.status_rules import _auto_status

def _vn_normalize(text: str) -> str:
    """
    Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
    Vai tro backend: Ham `_vn_normalize` la helper noi bo cua lop API trong file `api/views/bulk_upload.py`, chiu trach nhiem chuan hoa du lieu dau vao hoac du lieu trung gian truoc khi endpoint chinh phan hoi.
    Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can chuan hoa du lieu dau vao hoac du lieu trung gian nhung khong nen tu xu ly o client.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Thuong duoc cac ham public nhu `bulk_parse_excel`, `bulk_upload_single`, `template_replace_docx` goi lai.
    Tac dung: Co lap rieng buoc chuan hoa du lieu dau vao hoac du lieu trung gian de cac endpoint cung file tai su dung dung mot quy tac.
    """
    nfkd = unicodedata.normalize('NFKD', text)
    ascii_str = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return ascii_str.lower().strip()

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_parse_excel(request):
    """
    Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
    Vai tro backend: Ham `bulk_parse_excel` la endpoint hoac diem vao REST cua file `api/views/bulk_upload.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can thuc hien phan xu ly chuyen trach cua symbol hien tai tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Dung cung cap voi cac ham `_vn_normalize`, `bulk_upload_single`, `template_replace_docx` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac thuc hien phan xu ly chuyen trach cua symbol hien tai tren giao dien.
    """
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({'detail': 'Cần file excel_file.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active

        
        headers = []
        headers_norm = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            val = str(cell.value or '').strip()
            headers.append(val)
            headers_norm.append(_vn_normalize(val))

        

        def _col(candidates):
            """
            Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
            Vai tro backend: Ham `_col` la helper noi bo cua lop API trong file `api/views/bulk_upload.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
            Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
            Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Thuong duoc cac ham public nhu `bulk_parse_excel`, `bulk_upload_single`, `template_replace_docx` goi lai.
            Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
            """
            norm_cands = [_vn_normalize(c) for c in candidates]
            for nc in norm_cands:
                for i, h in enumerate(headers_norm):
                    if nc in h:
                        return i
            return None

        idx_filename    = _col(['ten file', 'filename', 'file'])
        idx_title       = _col(['tieu de mau', 'tieu de', 'ten mau', 'title'])
        idx_description = _col(['mo ta', 'description', 'ghi chu'])
        idx_eff         = _col(['ngay hieu luc', 'hieu luc', 'effective', 'ngay bat dau'])
        idx_end         = _col(['ngay het han', 'het han', 'end date', 'ngay ket thuc'])
        idx_tags        = _col(['tag', 'tags', 'tu khoa'])
        idx_groups      = _col([
            'nhom duoc phan mau', 'nhom phan phoi', 'nhom chia se',
            'phan phoi nhom', 'nhom', 'groups', 'group',
        ])

        rows = []
        for row in ws.iter_rows(min_row=2):
            vals = [str(cell.value or '').strip() for cell in row]
            if not any(vals):
                continue

            

            def _get(idx):
                """
                Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
                Vai tro backend: Ham `_get` la helper noi bo cua lop API trong file `api/views/bulk_upload.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
                Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
                Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Thuong duoc cac ham public nhu `bulk_parse_excel`, `bulk_upload_single`, `template_replace_docx` goi lai.
                Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
                """
                if idx is not None and idx < len(vals):
                    return vals[idx]
                return ''

            

            def _get_raw(idx):
                """
                Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
                Vai tro backend: Ham `_get_raw` la helper noi bo cua lop API trong file `api/views/bulk_upload.py`, chiu trach nhiem thuc hien phan xu ly chuyen trach cua symbol hien tai truoc khi endpoint chinh phan hoi.
                Vai tro cua no trong frontend: Frontend Flutter khong goi truc tiep helper nay; endpoint cung file dung no khi giao dien can thuc hien phan xu ly chuyen trach cua symbol hien tai nhung khong nen tu xu ly o client.
                Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Thuong duoc cac ham public nhu `bulk_parse_excel`, `bulk_upload_single`, `template_replace_docx` goi lai.
                Tac dung: Co lap rieng buoc thuc hien phan xu ly chuyen trach cua symbol hien tai de cac endpoint cung file tai su dung dung mot quy tac.
                """
                if idx is not None and idx < len(row):
                    return row[idx].value
                return None

            filename = _get(idx_filename)
            title = _get(idx_title)
            description = _get(idx_description)
            effective_date_raw = _get_raw(idx_eff)
            end_date_raw = _get_raw(idx_end)
            raw_tags = _get(idx_tags)
            raw_groups = _get(idx_groups)


            tags = [t.strip() for t in _re.split(r'[,;|]', raw_tags) if t.strip()] if raw_tags else []
            groups = [g.strip() for g in _re.split(r'[,;|]', raw_groups) if g.strip()] if raw_groups else []

            
            

            def _normalise_date(val):
                """Chuan hoa nhieu kieu input ve YYYY-MM-DD; tra '' neu khong
                hop le (vd "33/12/2026" hoac chuoi rac) thay vi day nguyen
                xuong DB de Django crash.
                """
                import datetime as _dt

                def _build(y, m_, d_):
                    try:
                        return _dt.date(int(y), int(m_), int(d_)).strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        return ''

                if val is None:
                    return ''
                if isinstance(val, (_dt.datetime, _dt.date)):
                    return val.strftime('%Y-%m-%d')
                s = str(val).strip()
                if not s or s.lower() in ('none', 'nan', ''):
                    return ''

                m = _re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s)
                if m:
                    return _build(m.group(3), m.group(2), m.group(1))

                m = _re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
                if m:
                    return _build(m.group(1), m.group(2), m.group(3))

                m = _re.match(r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$', s)
                if m:
                    return _build(m.group(1), m.group(2), m.group(3))

                try:
                    serial = float(s)
                    base = _dt.date(1899, 12, 30)
                    d = base + _dt.timedelta(days=int(serial))
                    return d.strftime('%Y-%m-%d')
                except (ValueError, OverflowError):
                    pass
                return ''

            rows.append({
                'filename': filename,
                'title': title,
                'description': description,
                'effective_date': _normalise_date(effective_date_raw),
                'end_date': _normalise_date(end_date_raw),
                'tags': tags,
                'groups': groups,
            })

        wb.close()
        return Response({'rows': rows})

    except Exception as e:
        return Response({'detail': f'Lỗi đọc file Excel: {e}'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_upload_single(request):
    """
    Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
    Vai tro backend: Ham `bulk_upload_single` la endpoint hoac diem vao REST cua file `api/views/bulk_upload.py`, chiu trach nhiem nhan tep tu frontend va luu xuong backend theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can nhan tep tu frontend va luu xuong backend tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Dung cung cap voi cac ham `_vn_normalize`, `bulk_parse_excel`, `bulk_upload_single` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac nhan tep tu frontend va luu xuong backend tren giao dien.
    """
    import base64

    from django.core.files.base import ContentFile
    from accounts.models import UserGroup
    from accounts.tenancy import get_user_company
    from document_templates.models import DocumentTemplate
    from document_templates.versioning import create_template_version_snapshot
    from .templates import _extract_docx_template_payload, _resolve_detection_guidance

    docx_file = request.FILES.get('docx_file')
    if not docx_file:
        return Response({'detail': 'Can docx_file.'}, status=status.HTTP_400_BAD_REQUEST)

    # Goi y nhan dien bien (tuy chon) tu nguoi dung — da chong injection + boc untrusted.
    guidance_block, guidance_error = _resolve_detection_guidance(request, request.user)
    if guidance_error is not None:
        return guidance_error

    import datetime as _dt

    def _safe_date(raw):
        """Tra date object hop le hoac None; tu choi chuoi nhu '2026-12-33'."""
        if not raw:
            return None
        s = str(raw).strip()
        if not s:
            return None
        m = _re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
        if not m:
            return None
        try:
            return _dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except (ValueError, TypeError):
            return None

    title = request.data.get('title', '').strip() or docx_file.name.rsplit('.', 1)[0]
    description = request.data.get('description', '').strip()
    effective_date_raw = request.data.get('effective_date', '').strip()
    end_date_raw = request.data.get('end_date', '').strip()
    effective_date = _safe_date(effective_date_raw)
    end_date = _safe_date(end_date_raw)
    invalid_dates = []
    if effective_date_raw and effective_date is None:
        invalid_dates.append(('effective_date', effective_date_raw))
    if end_date_raw and end_date is None:
        invalid_dates.append(('end_date', end_date_raw))
    generate_tags_flag = request.data.get('generate_tags', 'false').lower() == 'true'
    upload_mode = request.data.get('upload_mode', 'ai_detect').strip().lower() or 'ai_detect'

    if upload_mode not in {'ai_detect', 'prebuilt_variables'}:
        return Response(
            {'detail': 'upload_mode khong hop le.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        tags = request.data.get('tags', '[]')
        tags = _json.loads(tags) if isinstance(tags, str) else tags
        if not isinstance(tags, list):
            tags = []
    except Exception:
        tags = []
    tags = list(dict.fromkeys(str(tag).strip() for tag in tags if str(tag).strip()))

    try:
        raw_groups = request.data.get('groups', '[]')
        raw_groups = _json.loads(raw_groups) if isinstance(raw_groups, str) else raw_groups
        if not isinstance(raw_groups, list):
            raw_groups = []
    except Exception:
        raw_groups = []
    requested_group_names = list(dict.fromkeys(
        str(g).strip() for g in raw_groups if str(g).strip()
    ))

    if effective_date and end_date and end_date < effective_date:
        return Response(
            {'detail': 'Ngay het han khong duoc som hon ngay hieu luc.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        docx_bytes = docx_file.read()
        payload = _extract_docx_template_payload(
            docx_bytes,
            source_name=docx_file.name,
            auto_detect=(upload_mode == 'ai_detect'),
            guidance_block=guidance_block,
        )
    except Exception as exc:
        return Response({'detail': f'Loi doc DOCX: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    content_text = payload.get('content', '') or ''
    detected_vars = payload.get('detected_vars', []) or []
    modified_docx_b64 = payload.get('modified_docx') if upload_mode == 'ai_detect' else None
    persisted_docx_bytes = docx_bytes
    if isinstance(modified_docx_b64, str) and modified_docx_b64.strip():
        try:
            persisted_docx_bytes = base64.b64decode(modified_docx_b64)
        except Exception:
            persisted_docx_bytes = docx_bytes

    if upload_mode == 'prebuilt_variables' and not detected_vars:
        return Response(
            {'detail': 'File khong chua placeholder {{bien}} hop le.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if generate_tags_flag and not tags:
        try:
            from ai_engine.rag_engine import get_llm
            from langchain_core.messages import HumanMessage, SystemMessage

            llm = get_llm()
            tag_resp = llm.invoke([
                SystemMessage(content=(
                    "Ban la chuyen gia phan tich van ban hanh chinh. "
                    "Sinh 5-10 tags/tukhoa ngan gon phu hop de tim kiem ve sau. "
                    'Tra ve JSON: {"tags": ["tag1", ...]}.'
                )),
                HumanMessage(content=f"Tieu de: {title}\n\nNoi dung:\n{content_text[:2500]}"),
            ])
            match = _re.search(r'\{.*\}', str(tag_resp.content or '').strip(), _re.DOTALL)
            if match:
                parsed = _json.loads(match.group())
                tags = list(dict.fromkeys(
                    str(tag).strip() for tag in parsed.get('tags', []) if str(tag).strip()
                ))
        except Exception:
            pass

    user = request.user
    safe_name = _re.sub(r'[^\w.-]', '_', docx_file.name)

    # Resolve requested group names within the user's company. Matching is
    # both case-insensitive AND diacritics-insensitive so admin can type
    # either "Phòng Kế Toán" (chính xác) or "phong ke toan" (không dấu).
    # "đ"/"Đ" cũng được coi tương đương "d"/"D".
    def _group_key(s):
        nfkd = unicodedata.normalize('NFKD', s or '')
        ascii_only = ''.join(c for c in nfkd if not unicodedata.combining(c))
        return ascii_only.replace('đ', 'd').replace('Đ', 'D').lower().strip()

    matched_groups = []
    unmatched_names = []
    if requested_group_names:
        company = get_user_company(user)
        group_qs = UserGroup.objects.all()
        if company is not None:
            group_qs = group_qs.filter(company=company)
        existing = list(group_qs)
        # Build lookup keyed by both raw-lowercased name (preserves dấu)
        # and the fully-normalised key. Raw lookup wins so an exact match
        # never gets clobbered by a fuzzy collision.
        by_raw = {}
        by_norm = {}
        for g in existing:
            by_raw.setdefault(g.name.strip().lower(), g)
            by_norm.setdefault(_group_key(g.name), g)
        seen_ids = set()
        for name in requested_group_names:
            g = by_raw.get(name.strip().lower()) or by_norm.get(_group_key(name))
            if g is None:
                unmatched_names.append(name)
                continue
            if g.pk in seen_ids:
                continue
            seen_ids.add(g.pk)
            matched_groups.append(g)

    def _create_template(group=None):
        if group is not None:
            visibility = DocumentTemplate.VISIBILITY_GROUP
        else:
            visibility = DocumentTemplate.VISIBILITY_PRIVATE
        tstatus = _auto_status(DocumentTemplate.SOURCE_DOCX, visibility, user, group)
        t = DocumentTemplate.objects.create(
            owner=user,
            title=title,
            description=description,
            content=content_text,
            source_type=DocumentTemplate.SOURCE_DOCX,
            status=tstatus,
            visibility=visibility,
            group=group,
            tags=tags,
            effective_date=effective_date or None,
            end_date=end_date or None,
        )
        t.docx_file.save(safe_name, ContentFile(persisted_docx_bytes), save=False)
        t.save()
        create_template_version_snapshot(
            t,
            created_by=user,
            change_note=(
                f'Tao tu bulk upload (nhom: {group.name})'
                if group is not None else 'Tao tu bulk upload'
            ),
        )
        return t

    if matched_groups:
        created_templates = [_create_template(g) for g in matched_groups]
    else:
        created_templates = [_create_template(None)]
    tmpl = created_templates[0]

    # Ghi nhan tac vu tao mau vao khu vuc "tac vu" (AITaskProgress terminal) de
    # nguoi dung theo doi duoc tien trinh upload nhieu mau, dong nhat voi cac
    # tac vu AI khac. Khong duoc lam vo luong upload neu ghi task that bai.
    try:
        from django.utils import timezone as _tz

        from ai_tasks.models import (
            AITaskProgress,
            STATUS_COMPLETED,
            TASK_TYPE_BULK_TEMPLATE_UPLOAD,
        )

        AITaskProgress.objects.create(
            user=user,
            task_type=TASK_TYPE_BULK_TEMPLATE_UPLOAD,
            status=STATUS_COMPLETED,
            progress_percent=100,
            progress_stage='done',
            progress_detail=f'Da tao mau "{tmpl.title}".',
            title_summary=f'Tao mau: {tmpl.title}'[:255],
            deeplink=f'/templates/{tmpl.id}',
            related_entity_type='template',
            related_entity_id=tmpl.id,
            completed_at=_tz.now(),
            result={
                'created_ids': [t.id for t in created_templates],
                'assigned_groups': [g.name for g in matched_groups],
            },
        )
    except Exception:
        pass

    return Response({
        'id': tmpl.id,
        'title': tmpl.title,
        'tags': tmpl.tags,
        'detected_vars': detected_vars,
        'status': 'ok',
        'upload_mode': upload_mode,
        'created_ids': [t.id for t in created_templates],
        'assigned_groups': [g.name for g in matched_groups],
        'unmatched_groups': unmatched_names,
        'invalid_dates': [
            {'field': field, 'value': value} for field, value in invalid_dates
        ],
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def template_replace_docx(request, pk):
    """
    Thuoc chuc nang nao: Tao mau van ban, Mau dung chung, Mau phong ban, Mau rieng, Mau yeu thich va Tat ca mau (Admin).
    Vai tro backend: Ham `template_replace_docx` la endpoint hoac diem vao REST cua file `api/views/bulk_upload.py`, chiu trach nhiem xu ly du lieu hoac thao tac lien quan toi mau van ban theo request hien tai.
    Vai tro cua no trong frontend: Frontend goi truc tiep endpoint nay khi nguoi dung can xu ly du lieu hoac thao tac lien quan toi mau van ban tu man hinh tuong ung.
    Moi lien he voi nhung ham / source khac: Tuong tac truc tiep voi `api/urls.py`, `api/serializers/templates.py`, `accounts.permissions`, `document_templates.models`, `document_templates.utils`, `document_templates.versioning`. Dung cung cap voi cac ham `_vn_normalize`, `bulk_parse_excel`, `bulk_upload_single` trong module nay.
    Tac dung: Bien request thanh JSON, file hoac ma trang thai dung cho thao tac xu ly du lieu hoac thao tac lien quan toi mau van ban tren giao dien.
    """
    import base64

    from django.core.files.base import ContentFile
    from accounts.permissions import get_accessible_templates
    from document_templates.models import DocumentTemplate
    from document_templates.versioning import create_template_version_snapshot
    from .templates import _extract_docx_template_payload, _resolve_detection_guidance

    qs = get_accessible_templates(request.user)
    try:
        tmpl = qs.get(pk=pk)
    except DocumentTemplate.DoesNotExist:
        return Response({'detail': 'Khong tim thay mau.'}, status=status.HTTP_404_NOT_FOUND)

    if tmpl.owner != request.user and not request.user.is_superuser:
        return Response({'detail': 'Khong co quyen chinh sua.'}, status=status.HTTP_403_FORBIDDEN)

    docx_file = request.FILES.get('docx_file')
    if not docx_file:
        return Response({'detail': 'Can docx_file.'}, status=status.HTTP_400_BAD_REQUEST)

    change_note = request.data.get('change_note', '').strip() or f'Thay bang file DOCX moi: {docx_file.name}'

    guidance_block, guidance_error = _resolve_detection_guidance(request, request.user)
    if guidance_error is not None:
        return guidance_error

    try:
        docx_bytes = docx_file.read()
        payload = _extract_docx_template_payload(
            docx_bytes,
            source_name=docx_file.name,
            auto_detect=True,
            guidance_block=guidance_block,
        )
    except Exception as exc:
        return Response({'detail': f'Loi doc DOCX: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    create_template_version_snapshot(
        tmpl,
        created_by=request.user,
        change_note=change_note,
    )

    try:
        parts = tmpl.version.split('.')
        major = parts[0]
        minor = int(parts[1]) + 1 if len(parts) > 1 else 1
        tmpl.version = f'{major}.{minor}'
    except Exception:
        pass

    tmpl.content = payload.get('content', '') or ''
    tmpl.source_type = DocumentTemplate.SOURCE_DOCX
    tmpl.status = _auto_status(tmpl.source_type, tmpl.visibility, request.user, tmpl.group)
    if tmpl.status != 'approved':
        tmpl.approved_by = None
        tmpl.approved_at = None
        tmpl.approver_note = ''
    safe_name = _re.sub(r'[^\w.-]', '_', docx_file.name)
    modified_docx_b64 = payload.get('modified_docx')
    persisted_docx_bytes = docx_bytes
    if isinstance(modified_docx_b64, str) and modified_docx_b64.strip():
        try:
            persisted_docx_bytes = base64.b64decode(modified_docx_b64)
        except Exception:
            persisted_docx_bytes = docx_bytes
    tmpl.docx_file.save(safe_name, ContentFile(persisted_docx_bytes), save=False)
    tmpl.save()

    return Response({
        'id': tmpl.id,
        'content': tmpl.content,
        'detected_vars': payload.get('detected_vars', []) or [],
        'version': tmpl.version,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bulk_upload_sample(request):
    """Tra ve mot file upload.xlsx mau co san day du cot va vai dong vi du.

    Header giu nguyen kieu tieng Viet co dau de match logic doc o
    `bulk_parse_excel`. Cot "nhom duoc phan mau" co the de trong -- khi
    do template se theo luong cu (private, owner=admin).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mau'

    headers = [
        'ten file',
        'tieu de mau',
        'mo ta',
        'ngay hieu luc',
        'ngay het han',
        'tag',
        'nhom duoc phan mau',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sample_rows = [
        [
            'hop_dong_lao_dong.docx',
            'Hop dong lao dong',
            'Mau hop dong lao dong tieu chuan',
            '2025-01-01',
            '2026-12-31',
            'hop dong, nhan su',
            'Phong Nhan Su, Phong Hanh Chinh',
        ],
        [
            'quyet_dinh_tuyen_dung.docx',
            'Quyet dinh tuyen dung',
            'Mau quyet dinh tuyen dung nhan vien moi',
            '2025-01-01',
            '',
            'tuyen dung, quyet dinh',
            'Phong Nhan Su',
        ],
        [
            'mau_rieng_admin.docx',
            'Mau rieng cua admin',
            'De trong cot nhom -> chi luu vao "Mau cua toi" cua admin',
            '',
            '',
            '',
            '',
        ],
    ]
    for r in sample_rows:
        ws.append(r)

    widths = [28, 26, 38, 14, 14, 22, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
            value='Ghi chu cot "nhom duoc phan mau":\n'
                  '  - Tach nhieu nhom bang dau , ; hoac |\n'
                  '  - Ten nhom co the viet co dau ("Phong Ke Toan") hoac khong dau ("phong ke toan"), khong phan biet hoa-thuong\n'
                  '  - De trong -> template se la "private" thuoc admin (luong cu)').font = Font(italic=True, color='6B7280')
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[note_row].height = 60
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="upload.xlsx"'
    return response
